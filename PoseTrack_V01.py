from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import *
from PyQt5.QtWidgets import QApplication, QMainWindow, QVBoxLayout, QWidget, QFileDialog
from PyQt5.QtCore import QDate, QTime, QDateTime, Qt, QTimer
from PyQt5.QtGui import QPixmap, QImage
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from mpl_toolkits.mplot3d import Axes3D
from matplotlib.animation import FuncAnimation
import sys
import os
from os import remove
import cv2
import numpy as np
import glob
import pickle
import csv
from collections import deque
import matplotlib.pyplot as plt
import random
import copy
from scipy.interpolate import interp1d
import math
import pandas as pd
from subprocess import call
import json
from matplotlib import animation
from scipy import signal
import mediapipe as mp
import imageio
import time
import pandas as pd
import openpyxl
import csv
from scipy.spatial.transform import Rotation
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from scipy.signal import detrend



def list_files_with_prefix(directory_path, prefix="P_"):
    """
    List all files in the specified directory that start with the given prefix.

    :param directory_path: Path to the directory
    :param prefix: Prefix to filter file names
    :return: List of file names in the directory with the specified prefix
    """

    # Check if the given path is a valid directory
    if not os.path.isdir(directory_path):
        raise ValueError(f"'{directory_path}' is not a directory")

    # List all files in the directory with the given prefix
    return [f for f in os.listdir(directory_path) if
            os.path.isfile(os.path.join(directory_path, f)) and f.startswith(prefix)]

def list_files_without_prefix(directory_path):
    """
    List all files in the specified directory.
    :param directory_path: Path to the directory
    :return: List of file names in the directory with the specified prefix
    """

    # Check if the given path is a valid directory
    if not os.path.isdir(directory_path):
        raise ValueError(f"'{directory_path}' is not a directory")

    # List all files in the directory with the given prefix
    return [f for f in os.listdir(directory_path) if
            os.path.isfile(os.path.join(directory_path, f))]

def preprocessing_data_mediapipe(file_path):
    with open(file_path, "rb") as f:
        Cam_keypoints_data = pickle.load(f)

    # for key, value in Cam_keypoints_data.items():
    #     print(key, len(value))

    return Cam_keypoints_data

def DLT(P1, P2, point1, point2):
    A = [point1[1] * P1[2, :] - P1[1, :],
         P1[0, :] - point1[0] * P1[2, :],
         point2[1] * P2[2, :] - P2[1, :],
         P2[0, :] - point2[0] * P2[2, :]
         ]
    A = np.array(A).reshape((4, 4))
    B = A.transpose() @ A
    from scipy import linalg
    U, s, Vh = linalg.svd(B, full_matrices=False)
    return Vh[3, 0:3] / Vh[3, 3]

def three_dimensional_reconstruction(task_file_path_list, projection_file_list):
    for i in range(1, 6):
        for j in range(i + 1, 7):
            print("=*" * 30)
            # loading task keypoints in both cameras views
            for task_file_path in task_file_path_list:
                task_name = task_file_path.split('_')[1]

                # if task_file_path == f'cam{i}_{task_name}_task_keypoints_data.pkl':
                if task_file_path == f'Cam{i}_keypoints_data.pkl':
                    print("task file 1: ", task_file_path)
                    task_file_path_1 = os.path.join(tasks_pickle_files_directory, task_file_path)
                    Cam1_keypoints_data = preprocessing_data_mediapipe(task_file_path_1)
                    # print(Cam1_keypoints_data)

                # elif task_file_path == f'cam{j}_{task_name}_task_keypoints_data.pkl':
                elif task_file_path == f'Cam{j}_keypoints_data.pkl':
                    print("task file 2: ", task_file_path)
                    task_file_path_2 = os.path.join(tasks_pickle_files_directory, task_file_path)
                    Cam2_keypoints_data = preprocessing_data_mediapipe(task_file_path_2)
                    # print(Cam2_keypoints_data)

            # loading projections for both cameras
            for f1 in projection_file_list:
                # if f1 == f'P_cam{i}_cb2_(cam{i}_cb2-cam{j}_cb2).pkl':
                if f1 == f'P_CB{i}_(CB{i}-CB{j}).pkl':
                    print("projection matrix 1: ", f1)
                    projection_file_path_1 = os.path.join(projection_matrices_directory, f1)
                    with open(projection_file_path_1, "rb") as fp1:
                        P1 = pickle.load(fp1)

            for f2 in projection_file_list:
                # if f2 == f'P_cam{j}_cb2_(cam{i}_cb2-cam{j}_cb2).pkl':
                if f2 == f'P_CB{j}_(CB{i}-CB{j}).pkl':
                    print("projection matrix 2: ", f2)
                    projection_file_path_2 = os.path.join(projection_matrices_directory, f2)
                    with open(projection_file_path_2, "rb") as fp2:
                        P2 = pickle.load(fp2)

            keypoints_3D = []

            for key in Cam1_keypoints_data.keys():
                cam1_data_dict = Cam1_keypoints_data[key]
                cam2_data_dict = Cam2_keypoints_data[key]

                for data_1, data_2 in zip(cam1_data_dict, cam2_data_dict):
                    frame_number_1 = data_1['frame']
                    frame_number_2 = data_2['frame']
                    assert frame_number_1 == frame_number_2
                    point1 = data_1['coordinates']
                    point2 = data_2['coordinates']

                    p3d = DLT(P1, P2, point1, point2)
                    keypoints_3D.append(p3d)

            keypoints_3D = np.array(keypoints_3D)

            number_of_segments = len(Cam1_keypoints_data)
            number_of_video_frames = int(keypoints_3D.shape[0] / number_of_segments)
            keypoints_3D = np.reshape(keypoints_3D, (number_of_segments, number_of_video_frames, 3))

            master_dictionary[(i, j)] = keypoints_3D

    min_length = None
    for key, value in master_dictionary.items():
        length = value.shape[1]
        if min_length is None or length < min_length:
            min_length = length

    # Create a new dictionary to store trimmed arrays
    trimmed_dictionary = {}

    for key, value in master_dictionary.items():
        trimmed_array = value[:, :min_length, :]
        trimmed_dictionary[key] = trimmed_array

    print("=*"*30)
    for key, value in trimmed_dictionary.items():
        print(value.shape)
        print("Camera pair", key )
        # print("Reconstructed points : ", value)
        print("=*"*30)

    # Initialize the new dictionary to remove outliers
    transformed_dictionary = {}
    # Iterate over the keys in the master_dictionary
    for key, keypoints_3D in trimmed_dictionary.items():
        i, j = key
        # Iterate over the body joints
        for joint_number in range(keypoints_3D.shape[0]):
            joint_data = keypoints_3D[joint_number]
            # If the joint number is not in the new dictionary, create an entry
            if joint_number not in transformed_dictionary:
                transformed_dictionary[joint_number] = []
            # Append the joint data for the current (i, j) pair to the new dictionary
            transformed_dictionary[joint_number].append(joint_data)

    print("=*" * 30)
    for key, value in transformed_dictionary.items():
        print(np.array(value).shape)
        print("Joint", key)
        # print("Reconstructed points : ", value)
        print("=*" * 30)

    return transformed_dictionary

def remove_outliers(data):
    threshold = 1.5
    Q1 = np.percentile(data, 25)
    Q3 = np.percentile(data, 75)
    IQR = Q3 - Q1
    lower_bound = Q1 - threshold * IQR
    upper_bound = Q3 + threshold * IQR
    mid_point = (lower_bound + upper_bound) / 2
    return [x if lower_bound <= x <= upper_bound else mid_point for x in data]

def data_centering_and_outlier_removal(transformed_dictionary):
    centralized_series = {}

    # Loop through each key and its corresponding data array
    for key, data_list in transformed_dictionary.items():
        data_array = np.array(data_list)  # Convert list to NumPy array
        print(data_array.shape)

        keypoints_list = []

        # Loop through the dimensions and accumulate angle values
        for instance in range(data_array.shape[0]):
            keypoints_without_outliers_list = []
            for step in range(data_array.shape[1]):
                keypoints = data_array[instance, step, :]

                # Apply the remove_outliers function
                keypoints_without_outliers = remove_outliers(keypoints)

                # Append the angles_without_outliers to angles_without_outliers_list
                keypoints_without_outliers_list.append(keypoints_without_outliers)

            # Append angles_without_outliers_list to angles_list
            keypoints_list.append(keypoints_without_outliers_list)

        # Store the average angles in the dictionary
        centralized_series[key] = keypoints_list

    print("=*"*30)
    for key, value in centralized_series.items():
        print(np.array(value).shape)
        print("Joint", key )
        print("Reconstructed points : ", np.array(value))
        print("=*"*30)

    return centralized_series

def plot_centralized_series(centralized_series, keypoint_names):
    # Loop through each key and its corresponding data array
    for key, value in centralized_series.items():
        value_array = np.array(value)

        # Create a new figure for each key
        plt.figure(figsize=(12, 10))
        plt.suptitle(keypoint_names[key])  # Set the figure title

        # Loop through the dimensions for each subplot
        for dim in range(value_array.shape[2]):
            plt.subplot(3, 1, dim + 1)

            # Loop through each instance and plot its line
            for instance in range(value_array.shape[0]):
                plt.plot(value_array[instance, :, dim])

        plt.tight_layout()
        plt.show()

def data_averaging_procedure(centralized_series):
    average_data_dict = {}  # Dictionary to store average data for each key

    # Loop through each key and its corresponding data array
    for key, value in centralized_series.items():
        value_array = np.array(value)

        # Calculate the average along the instances dimension
        average_data = np.mean(value_array, axis=0)

        average_data_dict[key] = average_data  # Store average data in the dictionary

    ### Low pass Filtering ####
    b1, a1 = signal.butter(1, 2 / 7.5, 'low')
    average_data_filtered = {}
    for key, value in average_data_dict.items():
        average_values_filtered = signal.filtfilt(b1, a1, value, axis=0)
        average_data_filtered[key] = average_values_filtered

    return average_data_filtered

def plot_centralized_series_with_average(centralized_series, keypoint_names):

    average_data_dict = {}  # Dictionary to store average data for each key

    # Loop through each key and its corresponding data array
    for key, value in centralized_series.items():
        value_array = np.array(value)

        # Calculate the average along the instances dimension
        average_data = np.mean(value_array, axis=0)

        average_data_dict[key] = average_data  # Store average data in the dictionary

        # Create a new figure for each key
        plt.figure(figsize=(12, 10))
        plt.suptitle(keypoint_names[key])  # Set the figure title

        # Loop through the dimensions for each subplot
        for dim in range(value_array.shape[2]):
            plt.subplot(3, 1, dim + 1)

            # Loop through each instance and plot its line with reduced opacity
            for instance in range(value_array.shape[0]):
                plt.plot(value_array[instance, :, dim], color='gray', alpha=0.3)

            # Plot the averaged data with a solid black line
            plt.plot(average_data[:, dim], color='black')

        plt.tight_layout()
        plt.show()

    ### Low pass Filtering ####
    b1, a1 = signal.butter(1, 2 / 7.5, 'low')
    average_data_filtered = {}
    for key, value in average_data_dict.items():
        average_values_filtered = signal.filtfilt(b1, a1, value, axis=0)
        average_data_filtered[key] = average_values_filtered

    return average_data_filtered

def calculate_angle(a, b, c):
    # Convert input points to NumPy arrays
    a = np.array(a)  # First point
    b = np.array(b)  # Mid point
    c = np.array(c)  # End point

    # Calculate vectors from point b to points a and c
    vector1 = a - b
    vector2 = c - b

    # Calculate the dot product of the two vectors
    dot_product = np.dot(vector1, vector2)

    # Calculate the magnitudes of the vectors
    magnitude1 = np.linalg.norm(vector1)
    magnitude2 = np.linalg.norm(vector2)

    # Calculate the angle in radians
    radians = np.arccos(dot_product/(magnitude1*magnitude2))

    # Convert radians to degrees
    angle_degrees = radians * 180.0 / np.pi

    return angle_degrees

def calculate_absolute_angles(keypoints_coordinates_data, angle_dictionary):

    all_absolute_angles = {}

    for i in range (keypoints_coordinates_data.shape[1]):
        # Extract the x, y, and z coordinates for the points
        x = keypoints_coordinates_data[:, i, 0]
        y = keypoints_coordinates_data[:, i, 1]
        z = keypoints_coordinates_data[:, i, 2]

        # Get coordinates of keypoints for angle calculation
        points = np.stack([x, y, z], axis=-1)

        # The dictionary to store calculated absolute intersegmental angles
        absolute_angles = {}

        for angle_name, angle_pair in angle_dictionary.items():
            angle = calculate_angle(points[angle_pair[0]], points[angle_pair[1]], points[angle_pair[2]])
            absolute_angles[angle_name] = angle

        # Store the dictionary of angles for this `i` in the main dictionary
        all_absolute_angles[i] = absolute_angles

    def transform_dict(data):
        new_dict = {}
        for key, values in data.items():
            for sub_key, sub_value in values.items():
                # If the key isn't yet in the new_dict, initialize it with an empty list
                if sub_key not in new_dict:
                    new_dict[sub_key] = []
                new_dict[sub_key].append(sub_value)
        return new_dict

    # Transform each dictionary
    transformed_all_absolute_angles = transform_dict(all_absolute_angles)


    return transformed_all_absolute_angles, all_absolute_angles

def calculate_angle_between_vectors(v1, v2):
    """
    Compute the angle between two vectors, with angles in the range
    [0, 270) U (-90, 0].
    """
    # Compute the cosine of the angle
    cos_theta = np.dot(v1, v2) / (np.linalg.norm(v1) * np.linalg.norm(v2))

    # Compute the angle in degrees
    angle = np.arccos(cos_theta)
    angle = np.degrees(angle)

    # return int(np.ceil(angle))
    return angle

def calculate_local_axes_and_angles(points, segment_pair, global_x_axis, global_y_axis, global_z_axis):
    """
    Calculate the local coordinate axes and angles with respect to global coordinate axes.
    """
    # Calculate Y axis direction along the bone segment
    local_y_axis = (points[segment_pair[1]] - points[segment_pair[0]]) / np.linalg.norm(points[segment_pair[1]] - points[segment_pair[0]])

    # Calculate X axis as perpendicular to Y and global Z axis direction
    local_x_axis = np.cross(local_y_axis, global_z_axis)
    local_x_axis /= np.linalg.norm(local_x_axis)

    # Calculate Z axis as perpendicular to X and Y axis direction
    local_z_axis = np.cross(local_x_axis, local_y_axis)

    # Calculate angles between local and global axes
    angle_x = calculate_angle_between_vectors(local_x_axis, global_x_axis)
    angle_y = calculate_angle_between_vectors(local_y_axis, global_y_axis)
    angle_z = calculate_angle_between_vectors(local_z_axis, global_z_axis)

    return local_x_axis, local_y_axis, local_z_axis, angle_x, angle_y, angle_z

def calculate_extrinsic_segment_angles(keypoints_coordinates_data, segments_dictionary):

    global_x_axis = np.array([1, 0, 0])
    global_y_axis = np.array([0, 1, 0])
    global_z_axis = np.array([0, 0, 1])

    all_extrinsic_angles = {}

    for i in range(keypoints_coordinates_data.shape[1]):
        # Extract the x, y, and z coordinates for the points
        x = keypoints_coordinates_data[:, i, 0]
        y = keypoints_coordinates_data[:, i, 1]
        z = keypoints_coordinates_data[:, i, 2]

        # Get coordinates of keypoints for coordinate frame calculation
        points = np.stack([x, y, z], axis=-1)

        extrinsic_angles = {}

        for segment_name, segment_pair in segments_dictionary.items():
            local_x_axis, local_y_axis, local_z_axis, angle_x, angle_y, angle_z = calculate_local_axes_and_angles(points,
                                                                                                                  segment_pair,
                                                                                                                  global_x_axis,
                                                                                                                  global_y_axis,
                                                                                                                  global_z_axis)

            # extrinsic_angles[segment_name] = [int(np.ceil(angle_x)), int(np.ceil(angle_y)), int(np.ceil(angle_z))]
            extrinsic_angles[segment_name] = [angle_x, angle_y, angle_z]


        # Store the dictionary of angles for this `i` in the main dictionary
        all_extrinsic_angles[i] = extrinsic_angles

    return all_extrinsic_angles

def plot_segment_angles(segment_name, rotation_dict):
    # Extract the extrinsic angles for the given segment_name
    extrinsic_rot_x = [value[0] for value in rotation_dict.values()]
    extrinsic_rot_y = [value[1] for value in rotation_dict.values()]
    extrinsic_rot_z = [value[2] for value in rotation_dict.values()]

    fig, (ax1, ax2, ax3) = plt.subplots(3, 1, sharex=True, figsize=(10, 8))
    fig.suptitle(f"Rotation Angles for {segment_name}")

    # Plot the X rotation angles
    ax1.plot(extrinsic_rot_x, 'r-')
    ax1.set_ylabel("X Rotation Angle")

    # Plot the Y rotation angles
    ax2.plot(extrinsic_rot_y, 'g-')
    ax2.set_ylabel("Y Rotation Angle")

    # Plot the Z rotation angles
    ax3.plot(extrinsic_rot_z, 'b-')
    ax3.set_xlabel("Frame")
    ax3.set_ylabel("Z Rotation Angle")

    plt.tight_layout()
    plt.show()

def save_extrinsic_angles_to_excel(all_extrinsic_angles, filename):
    wb = openpyxl.Workbook()
    # Remove the default sheet that is created automatically
    default_sheet = wb.active
    wb.remove(default_sheet)

    for seg in segment_names:
        sheet = wb.create_sheet(title=seg)
        angles_data = all_extrinsic_angles.get(seg, {})  # Get angle data for the segment

        if angles_data:
            current_sheet = wb[seg]  # Get the sheet object
            current_sheet.append(["Frame", "X", "Y", "Z"])

            for frame, angles in angles_data.items():
                current_sheet.append([frame] + angles)  # Write angle values for each frame

    wb.save(filename)

def  animate_body_skeleton_wo_coordinate_frame(keypoints_coordinates_data, segment_pairs, joint_angles_frame_by_frame):

    number_of_animation_frames = keypoints_coordinates_data.shape[1]

    # Initialize body skeleton figure
    fig1 = plt.figure(figsize=(12, 8))
    ax1 = fig1.add_subplot(111, projection='3d')

    def animate(i):
        ax1.cla()

        x = keypoints_coordinates_data[:, i, 0]
        y = keypoints_coordinates_data[:, i, 1]
        z = keypoints_coordinates_data[:, i, 2]
        # Get coordinates of keypoints for angle calculation
        points = np.stack([x, y, z], axis=-1)

        # set the 3d plot limits
        # Compute the global minima and maxima
        x_min = np.min(keypoints_coordinates_data[:, :, 0])
        x_max = np.max(keypoints_coordinates_data[:, :, 0])

        y_min = np.min(keypoints_coordinates_data[:, :, 1])
        y_max = np.max(keypoints_coordinates_data[:, :, 1])

        z_min = np.min(keypoints_coordinates_data[:, :, 2])
        z_max = np.max(keypoints_coordinates_data[:, :, 2])
        ######
        # Plot joints in orange
        ax1.scatter(x, y, z, color='red')
        for m in range(keypoints_coordinates_data.shape[0]):
            ax1.text(x[m] - 0.05, y[m] - 0.05, z[m] - 0.05, f'{m}', color="orange", fontsize=9)

        # Create an empty dictionary to store legend labels and corresponding artists
        legend_dict = {}

        for joint_name, angle_value in joint_angles_frame_by_frame[i].items():
            angle_value = str(int(angle_value))  # Convert angle_value to a string
            joint_coords = points[list(joint_angles_frame_by_frame[i].keys()).index(joint_name)]

            # Create a label for the legend
            label = f'{joint_name}: {angle_value}'

            # Create a dummy scatter point (you can customize this)
            dummy_scatter = ax1.scatter([], [], [], label=label, color='gray', s=30)

            # Add the dummy scatter point to the legend dictionary with the label
            legend_dict[label] = dummy_scatter

        # Display the legend on the plot
        ax1.legend(legend_dict.values(), legend_dict.keys(), loc='best', fontsize=9)

        # Calculate global Z axis
        global_z_axis = np.array([0, 0, 1])  # Assuming global Z axis is [0, 0, 1]

        # Draw the bone segments in black
        for idx, (keypoint1, keypoint2) in enumerate(segment_pairs):

            # Calculate Y axis direction along the bone segment
            y_axis = (points[keypoint2] - points[keypoint1]) / np.linalg.norm(points[keypoint2] - points[keypoint1])

            # Calculate X axis direction perpendicular to Y and global Z axis direction
            x_axis = np.cross(y_axis, global_z_axis)
            x_axis /= np.linalg.norm(x_axis)

            # Calculate Z axis direction perpendicular to Y and X axis direction
            z_axis = np.cross(y_axis, x_axis)
            z_axis /= np.linalg.norm(z_axis)

            ax1.plot([points[keypoint1][0], points[keypoint2][0]], [points[keypoint1][1], points[keypoint2][1]],
                    [points[keypoint1][2], points[keypoint2][2]], color='black')


            ax1.set_xlim([x_min, x_max])
            ax1.set_ylim([y_min, y_max])
            ax1.set_zlim([z_min, z_max])

            # Set equal aspect ratio
            ax1.set_aspect('equal')

            ax1.invert_zaxis()
            ax1.invert_xaxis()


    def on_rotation(event):
        azim, elev = ax1.azim, ax1.elev
        print(f"Azimuth: {azim}, Elevation: {elev}")

    # Create animation
    # ani_1 = FuncAnimation(fig1, animate, frames=range(0,number_of_video_frames), interval=40)

    ani_1 = FuncAnimation(fig1, animate, frames=range(0, number_of_animation_frames), interval=33)
    # Set the view
    ax1.view_init(elev=110.65, azim=90.78)

    # # Save as .mp4 file
    ani_1.save(ProjectWorkingPath + '/task_result_files/animated_model_wo_coordinate_frames.mp4', writer='ffmpeg')

    # Connect the event to the function
    # c1 = fig1.canvas.mpl_connect('motion_notify_event', on_rotation)
    # plt.show()
    print("Animated model is saved!")

def animate_body_skeleton_with_coordinate_frame(keypoints_coordinates_data, segment_pairs):
    number_of_animation_frames = keypoints_coordinates_data.shape[1]

    # Initialize body skeleton figure
    fig1 = plt.figure(figsize=(12, 8))
    ax1 = fig1.add_subplot(111, projection='3d')

    def animate(i):

        ax1.cla()

        x = keypoints_coordinates_data[:, i, 0]
        y = keypoints_coordinates_data[:, i, 1]
        z = keypoints_coordinates_data[:, i, 2]
        # Get coordinates of keypoints for angle calculation
        points = np.stack([x, y, z], axis=-1)

        # set the 3d plot limits
        # Compute the global minima and maxima
        x_min = np.min(keypoints_coordinates_data[:, :, 0])
        x_max = np.max(keypoints_coordinates_data[:, :, 0])

        y_min = np.min(keypoints_coordinates_data[:, :, 1])
        y_max = np.max(keypoints_coordinates_data[:, :, 1])

        z_min = np.min(keypoints_coordinates_data[:, :, 2])
        z_max = np.max(keypoints_coordinates_data[:, :, 2])

        # Calculate the center of the scene
        scene_center = np.array([(x_min + x_max) / 2, (y_min + y_max) / 2, (z_min + z_max) / 2])

        ######
        # Plot joints in orange
        ax1.scatter(x, y, z, color='red')
        for m in range(keypoints_coordinates_data.shape[0]):
            ax1.text(x[m] - 0.05, y[m] - 0.05, z[m] - 0.05, f'{m}', color="orange", fontsize=9)

        # Calculate global Z axis
        global_z_axis = np.array([0, 0, 1])  # Assuming global Z axis is [0, 0, 1]

        # Draw the bone segments in black
        for idx, (keypoint1, keypoint2) in enumerate(segment_pairs):
            # Calculate the midpoint of the bone segment
            bone_midpoint = (points[keypoint1] + points[keypoint2]) / 2

            # Calculate Y axis direction along the bone segment
            y_axis = (points[keypoint2] - points[keypoint1]) / np.linalg.norm(points[keypoint2] - points[keypoint1])

            # Calculate X axis direction perpendicular to Y and global Z axis direction
            x_axis = np.cross(y_axis, global_z_axis)
            x_axis /= np.linalg.norm(x_axis)

            # Calculate Z axis direction perpendicular to Y and X axis direction
            z_axis = np.cross(y_axis, x_axis)
            z_axis /= np.linalg.norm(z_axis)

            ax1.plot([points[keypoint1][0], points[keypoint2][0]], [points[keypoint1][1], points[keypoint2][1]],
                     [points[keypoint1][2], points[keypoint2][2]], color='black')

            # Draw coordinate frames at the midpoint of the bone segment with increased linewidth
            # Set the fixed scale factor for the quiver lengths
            scale_factor = 0.1

            ax1.quiver(bone_midpoint[0], bone_midpoint[1], bone_midpoint[2], x_axis[0], x_axis[1], x_axis[2], color='r',
                      length=scale_factor, linewidth=2)
            ax1.text(bone_midpoint[0] + scale_factor * x_axis[0],
                    bone_midpoint[1] + scale_factor * x_axis[1],
                    bone_midpoint[2] + scale_factor * x_axis[2], 'X', color='r')

            ax1.quiver(bone_midpoint[0], bone_midpoint[1], bone_midpoint[2], y_axis[0], y_axis[1], y_axis[2], color='g',
                      length=scale_factor, linewidth=2)
            ax1.text(bone_midpoint[0] + scale_factor * y_axis[0],
                    bone_midpoint[1] + scale_factor * y_axis[1],
                    bone_midpoint[2] + scale_factor * y_axis[2], 'Y', color='g')

            ax1.quiver(bone_midpoint[0], bone_midpoint[1], bone_midpoint[2], z_axis[0], z_axis[1], z_axis[2], color='b',
                      length=scale_factor, linewidth=2)
            ax1.text(bone_midpoint[0] + scale_factor * z_axis[0],
                    bone_midpoint[1] + scale_factor * z_axis[1],
                    bone_midpoint[2] + scale_factor * z_axis[2], 'Z', color='b')

            # Plot global X, Y, and Z axes with dashed lines

            global_x = scene_center + scale_factor * 10 * np.array([1, 0, 0])
            global_y = scene_center + scale_factor * 10 * np.array([0, 1, 0])
            global_z = scene_center + scale_factor * 10 * np.array([0, 0, 1])

            # Plot dashed global X, Y, and Z axes
            ax1.plot([scene_center[0], global_x[0]], [scene_center[1], global_x[1]], [scene_center[2], global_x[2]],
                     linestyle = '--', linewidth = 1,  color = 'r')
            ax1.plot([scene_center[0], global_y[0]], [scene_center[1], global_y[1]], [scene_center[2], global_y[2]],
                     linestyle = '--', linewidth = 1, color = 'g')
            ax1.plot([scene_center[0], global_z[0]], [scene_center[1], global_z[1]], [scene_center[2], global_z[2]],
                     linestyle = '--', linewidth = 1, color = 'b')

            # # Define global axes positions
            # global_x = np.array([1, 0, 0])
            # global_y = np.array([0, 1, 0])
            # global_z = np.array([0, 0, 1])
            #
            # global_x = bone_midpoint + scale_factor * 2 * global_x
            # global_y = bone_midpoint + scale_factor * 2 * global_y
            # global_z = bone_midpoint + scale_factor * 2 * global_z
            #
            # ax1.plot([bone_midpoint[0], global_x[0]], [bone_midpoint[1], global_x[1]], [bone_midpoint[2], global_x[2]],
            #          linestyle='--', color='r')
            # ax1.plot([bone_midpoint[0], global_y[0]],[bone_midpoint[1], global_y[1]], [bone_midpoint[2], global_y[2]],
            #          linestyle = '--', color = 'g')
            # ax1.plot([bone_midpoint[0], global_z[0]],[bone_midpoint[1], global_z[1]], [bone_midpoint[2], global_z[2]],
            #          linestyle = '--', color = 'b')

            ax1.set_xlim([x_min, x_max])
            ax1.set_ylim([y_min, y_max])
            ax1.set_zlim([z_min, z_max])

            # Set equal aspect ratio
            ax1.set_aspect('equal')

            ax1.invert_zaxis()
            ax1.invert_xaxis()


    def on_rotation(event):
        azim, elev = ax1.azim, ax1.elev
        print(f"Azimuth: {azim}, Elevation: {elev}")

    # Create animation
    # ani_1 = FuncAnimation(fig1, animate, frames=range(0,number_of_video_frames), interval=40)

    ani_1 = FuncAnimation(fig1, animate, frames=range(0, number_of_animation_frames), interval=33)
    # Set the view
    ax1.view_init(elev=110.65, azim=90.78)

    # # Save as .mp4 file
    ani_1.save(ProjectWorkingPath + '/task_result_files/animated_model_with_coordinate_frames.mp4', writer='ffmpeg')

    # Connect the event to the function
    # c1 = fig1.canvas.mpl_connect('motion_notify_event', on_rotation)
    # plt.show()
    print("Animated model is saved!")

def calculate_intesegment_angles(segment_name_1, segment_name_2, all_extrinsic_angles, scale1, scale2):
    rotation_differences = {}    # Extract the extrinsic angles for the given segment names

    extrinsic_rot_x_segment_1 = [all_extrinsic_angles[i][segment_name_1][0]*scale1 for i in all_extrinsic_angles]
    extrinsic_rot_y_segment_1 = [all_extrinsic_angles[i][segment_name_1][1]*scale1 for i in all_extrinsic_angles]
    extrinsic_rot_z_segment_1 = [all_extrinsic_angles[i][segment_name_1][2]*scale1 for i in all_extrinsic_angles]

    extrinsic_rot_x_segment_2 = [all_extrinsic_angles[i][segment_name_2][0]*scale2 for i in all_extrinsic_angles]
    extrinsic_rot_y_segment_2 = [all_extrinsic_angles[i][segment_name_2][1]*scale2 for i in all_extrinsic_angles]
    extrinsic_rot_z_segment_2 = [all_extrinsic_angles[i][segment_name_2][2]*scale2 for i in all_extrinsic_angles]

    # Detrend the lists
    extrinsic_rot_x_segment_1 = detrend(extrinsic_rot_x_segment_1)
    extrinsic_rot_y_segment_1 = detrend(extrinsic_rot_y_segment_1)
    extrinsic_rot_z_segment_1 = detrend(extrinsic_rot_z_segment_1)

    extrinsic_rot_x_segment_2 = detrend(extrinsic_rot_x_segment_2)
    extrinsic_rot_y_segment_2 = detrend(extrinsic_rot_y_segment_2)
    extrinsic_rot_z_segment_2 = detrend(extrinsic_rot_z_segment_2)

    # Calculate the differences for each component (diagonal)
    diff_rot_x = [extrinsic_rot_x_segment_2[i] - extrinsic_rot_x_segment_1[i] for i in
                  range(len(extrinsic_rot_x_segment_1))]
    diff_rot_y = [extrinsic_rot_y_segment_2[i] - extrinsic_rot_y_segment_1[i] for i in
                  range(len(extrinsic_rot_y_segment_1))]
    diff_rot_z = [extrinsic_rot_z_segment_2[i] - extrinsic_rot_z_segment_1[i] for i in
                  range(len(extrinsic_rot_z_segment_1))]

    # Calculate x-y, x-z, and y-z angles (non-diagonal)
    diff_rot_x_y = [extrinsic_rot_x_segment_2[i] - extrinsic_rot_y_segment_1[i] for i in
                    range(len(extrinsic_rot_y_segment_1))]
    diff_rot_x_z = [extrinsic_rot_x_segment_2[i] - extrinsic_rot_z_segment_1[i] for i in
                    range(len(extrinsic_rot_z_segment_1))]
    diff_rot_y_z = [extrinsic_rot_y_segment_2[i] - extrinsic_rot_z_segment_1[i] for i in
                    range(len(extrinsic_rot_z_segment_1))]

    # Store the differences in the dictionary
    rotation_differences['diff_rot_x'] = diff_rot_x
    rotation_differences['diff_rot_y'] = diff_rot_y
    rotation_differences['diff_rot_z'] = diff_rot_z
    rotation_differences['diff_rot_x_y'] = diff_rot_x_y
    rotation_differences['diff_rot_x_z'] = diff_rot_x_z
    rotation_differences['diff_rot_y_z'] = diff_rot_y_z

    # Create a figure with subplots arranged in groups for diagonal angles
    fig1, axs1 = plt.subplots(3, 2, figsize=(12, 8))

    # Plot extrinsic_rot_x in the upper left subplot
    axs1[0, 0].plot(extrinsic_rot_x_segment_1, 'k', linestyle='solid', label=f'{segment_name_1} X')
    axs1[0, 0].plot(extrinsic_rot_x_segment_2, 'gray', linestyle='solid', label=f'{segment_name_2} X')
    axs1[0, 0].set_ylabel("Extrinsic Rotation X")
    axs1[0, 0].legend(loc='upper right')

    # Plot diff_rot_x in the lower left subplot
    axs1[0, 1].plot(diff_rot_x, 'r', label='X-X')
    axs1[0, 1].set_ylabel("Angle Difference")
    axs1[0, 1].legend(loc='upper right')

    # Plot extrinsic_rot_y in the upper right subplot
    axs1[1, 0].plot(extrinsic_rot_y_segment_1, 'k', linestyle='solid', label=f'{segment_name_1} Y')
    axs1[1, 0].plot(extrinsic_rot_y_segment_2, 'gray', linestyle='solid', label=f'{segment_name_2} Y')
    axs1[1, 0].set_ylabel("Extrinsic Rotation Y")
    axs1[1, 0].legend(loc='upper right')

    # Plot diff_rot_y in the lower right subplot
    axs1[1, 1].plot(diff_rot_y, 'g', label='Y-Y')
    axs1[1, 1].set_ylabel("Angle Difference")
    axs1[1, 1].legend(loc='upper right')

    # Plot extrinsic_rot_z in the upper middle subplot
    axs1[2, 0].plot(extrinsic_rot_z_segment_1, 'k', linestyle='solid', label=f'{segment_name_1} Z')
    axs1[2, 0].plot(extrinsic_rot_z_segment_2, 'gray', linestyle='solid', label=f'{segment_name_2} Z')
    axs1[2, 0].set_xlabel("Frames")
    axs1[2, 0].set_ylabel("Extrinsic Rotation Z")
    axs1[2, 0].legend(loc='upper right')

    # Plot diff_rot_z in the lower middle subplot
    axs1[2, 1].plot(diff_rot_z, 'b', label='Z-Z')
    axs1[2, 1].set_xlabel("Frames")
    axs1[2, 1].set_ylabel("Angle Difference")
    axs1[2, 1].legend(loc='upper right')

    # Adjust subplot spacing for the diagonal angles figure
    # plt.tight_layout()

    # Create a figure with subplots arranged in groups for non-diagonal angles
    fig2, axs2 = plt.subplots(3, 2, figsize=(12, 8))

    # Plot extrinsic_rot_x_y in the upper left subplot
    axs2[0, 0].plot(extrinsic_rot_x_segment_1, 'k', linestyle='solid', label=f'{segment_name_1} X')
    axs2[0, 0].plot(extrinsic_rot_y_segment_2, 'gray', linestyle='solid', label=f'{segment_name_2} Y')
    axs2[0, 0].set_ylabel("Extrinsic Rotation X and Y")
    axs2[0, 0].legend(loc='upper right')

    # Plot diff_rot_x_y in the lower left subplot
    axs2[0, 1].plot(diff_rot_x_y, 'c', label='X-Y')
    axs2[0, 1].set_ylabel("Angle Difference")
    axs2[0, 1].legend(loc='upper right')

    # Plot extrinsic_rot_x_z in the upper right subplot
    axs2[1, 0].plot(extrinsic_rot_x_segment_1, 'k', linestyle='solid', label=f'{segment_name_1} X')
    axs2[1, 0].plot(extrinsic_rot_z_segment_2, 'gray', linestyle='solid', label=f'{segment_name_2} Z')
    axs2[1, 0].set_ylabel("Extrinsic Rotation X and Z")
    axs2[1, 0].legend(loc='upper right')

    # Plot diff_rot_x_z in the lower right subplot
    axs2[1, 1].plot(diff_rot_x_z, 'm', label='X-Z')
    axs2[1, 1].set_ylabel("Angle Difference")
    axs2[1, 1].legend(loc='upper right')

    # Plot extrinsic_rot_y_z in the upper middle subplot
    axs2[2, 0].plot(extrinsic_rot_y_segment_1, 'k', linestyle='solid', label=f'{segment_name_1} Y')
    axs2[2, 0].plot(extrinsic_rot_z_segment_2, 'gray', linestyle='solid', label=f'{segment_name_2} Z')
    axs2[2, 0].set_xlabel("Frames")
    axs2[2, 0].set_ylabel("Extrinsic Rotation Y and Z")
    axs2[2, 0].legend(loc='upper right')

    # Plot diff_rot_y_z in the lower middle subplot
    axs2[2, 1].plot(diff_rot_y_z, 'y', label='Y-Z')
    axs2[2, 1].set_xlabel("Frames")
    axs2[2, 1].set_ylabel("Angle Difference")
    axs2[2, 1].legend(loc='upper right')

    # Adjust subplot spacing for the non-diagonal angles figure
    # plt.tight_layout()

    # Set titles and labels for both figures
    fig1.suptitle(f"Diagonal Intersegment Angles between {segment_name_1} and {segment_name_2}")
    fig2.suptitle(f"Non-Diagonal Intersegment Angles between {segment_name_1} and {segment_name_2}")

    # Show both figures with grouped subplots
    plt.show()
    return rotation_differences

def save_intrinsic_angles_to_excel(intrinsic_angle, filename):
    wb = openpyxl.Workbook()
    # Remove the default sheet that is created automatically
    default_sheet = wb.active
    wb.remove(default_sheet)

    direction_names = ["diff_rot_x", "diff_rot_y", "diff_rot_z", "diff_rot_x_y", "diff_rot_x_z", "diff_rot_y_z"]

    for direction in direction_names:
        sheet = wb.create_sheet(title=direction)
        intrinsic_angle_data = intrinsic_angle.get(direction, {})  # Get angle data for the key

        if intrinsic_angle_data:
            current_sheet = wb[direction]  # Get the sheet object
            for frame, angle_value in enumerate(intrinsic_angle_data, start=1):
                current_sheet.cell(row=frame, column=1, value=angle_value)  # Write angle values in a single column

    wb.save(filename)

def calculate_planar_angles(a, b, c):
    # Convert input points to NumPy arrays
    a = np.array(a)  # First point
    b = np.array(b)  # Mid point
    c = np.array(c)  # End point

    # Define your two vectors A and B
    A = [a[0] - b[0], a[1] - b[1], a[2] - b[2]]
    B = [c[0] - b[0], c[1] - b[1], c[2] - b[2]]

    # Calculate the dot product of A and B
    dot_product = sum([A[i] * B[i] for i in range(3)])

    # Calculate the magnitudes (lengths) of both vectors
    magnitude_A = math.sqrt(sum([A[i] ** 2 for i in range(3)]))
    magnitude_B = math.sqrt(sum([B[i] ** 2 for i in range(3)]))

    # Calculate the angles in each plane
    # Angle in the XY plane (θxy)
    theta_xy = math.acos((A[0] * B[0] + A[1] * B[1]) / (magnitude_A * magnitude_B))

    # Angle in the YZ plane (θyz)
    theta_yz = math.acos((A[1] * B[1] + A[2] * B[2]) / (magnitude_A * magnitude_B))

    # Angle in the XZ plane (θxz)
    theta_xz = math.acos((A[0] * B[0] + A[2] * B[2]) / (magnitude_A * magnitude_B))

    # Convert angles from radians to degrees if needed
    theta_xy_degrees = math.degrees(theta_xy)
    theta_yz_degrees = math.degrees(theta_yz)
    theta_xz_degrees = math.degrees(theta_xz)

    return theta_xy_degrees, theta_yz_degrees, theta_xz_degrees

def calculate_planar_joints_angles(keypoints_coordinates_data, angle_dictionary):

    all_planar_joints_angles = {}

    for i in range (keypoints_coordinates_data.shape[1]):
        # Extract the x, y, and z coordinates for the points
        x = keypoints_coordinates_data[:, i, 0]
        y = keypoints_coordinates_data[:, i, 1]
        z = keypoints_coordinates_data[:, i, 2]

        # Get coordinates of keypoints for angle calculation
        points = np.stack([x, y, z], axis=-1)

        # The dictionary to store calculated absolute intersegmental angles
        planar_joints_angles = {}

        for angle_name, angle_pair in angle_dictionary.items():
            angle_x, angle_y, angle_z = calculate_planar_angles(points[angle_pair[0]], points[angle_pair[1]], points[angle_pair[2]])
            planar_joints_angles[angle_name] = angle_x, angle_y, angle_z


        # Store the dictionary of angles for this `i` in the main dictionary
        all_planar_joints_angles[i] = planar_joints_angles

    def transform_dict(data):
        new_dict = {}
        for key, values in data.items():
            for sub_key, sub_value in values.items():
                # If the key isn't yet in the new_dict, initialize it with an empty list
                if sub_key not in new_dict:
                    new_dict[sub_key] = []
                new_dict[sub_key].append(sub_value)
        return new_dict

    # Transform each dictionary
    transformed_all_planar_joints_angles = transform_dict(all_planar_joints_angles)

    return transformed_all_planar_joints_angles

def calculate_euler_angles_from_segments_orientations(segment_name_1, segment_name_2, all_extrinsic_angles):
    # Extract the extrinsic angles for the given segment names and convert to radians
    extrinsic_orientations_segment_1 = [all_extrinsic_angles[i][segment_name_1] for i in all_extrinsic_angles]
    extrinsic_orientations_segment_2 = [all_extrinsic_angles[i][segment_name_2] for i in all_extrinsic_angles]

    segment_1_rad = np.radians(extrinsic_orientations_segment_1)
    segment_2_rad = np.radians(extrinsic_orientations_segment_2)

    # Initialize empty lists to store the relative Euler angles
    relative_euler_angles_roll = []
    relative_euler_angles_pitch = []
    relative_euler_angles_yaw = []

    for element1, element2 in zip(segment_1_rad, segment_2_rad):
        # Create rotation matrices for both segments
        R_segment_1 = np.dot(
            np.dot(
                np.array([[1, 0, 0],
                          [0, np.cos(element1[0]), -np.sin(element1[0])],
                          [0, np.sin(element1[0]), np.cos(element1[0])]]),
                np.array([[np.cos(element1[1]), 0, np.sin(element1[1])],
                          [0, 1, 0],
                          [-np.sin(element1[1]), 0, np.cos(element1[1])]])),
            np.array([[np.cos(element1[2]), -np.sin(element1[2]), 0],
                      [np.sin(element1[2]), np.cos(element1[2]), 0],
                      [0, 0, 1]]))

        R_segment_2 = np.dot(
            np.dot(
                np.array([[1, 0, 0],
                          [0, np.cos(element2[0]), -np.sin(element2[0])],
                          [0, np.sin(element2[0]), np.cos(element2[0])]]),
                np.array([[np.cos(element2[1]), 0, np.sin(element2[1])],
                          [0, 1, 0],
                          [-np.sin(element2[1]), 0, np.cos(element2[1])]])),
            np.array([[np.cos(element2[2]), -np.sin(element2[2]), 0],
                      [np.sin(element2[2]), np.cos(element2[2]), 0],
                      [0, 0, 1]]))

        # Calculate the rotation matrix to transform the lower leg to the upper leg
        R_relative = np.dot(R_segment_1.T, R_segment_2)

        r = Rotation.from_matrix(R_relative)

        # Convert radians to degrees
        euler_angles_deg = r.as_euler('xyz', degrees=True)

        relative_euler_angles_roll.append(euler_angles_deg[0])
        relative_euler_angles_pitch.append(euler_angles_deg[1])
        relative_euler_angles_yaw.append(euler_angles_deg[2])

    return relative_euler_angles_roll, relative_euler_angles_pitch, relative_euler_angles_yaw


### Important coding for keypoints ###
# right_shoulder : 0
# left_shoulder : 1
# mid_shoulder : 2
# right_wrist: 3
# left_wrist: 4
# right_elbow: 5
# left_elbow : 6
# right_hip : 7
# left_hip : 8
# mid_hip : 9
# right_knee : 10
# left_knee : 11
# right_ankle : 12
# left_ankle : 13
# head : 14
############################

# Define pairs of keypoints to connect
segment_pairs = [(0,1),
                 (2,14),
                 (0,5), (5,3),
                 (1,6), (6,4),
                 (7,8),
                 (2,9),
                 (7,10), (10,12),
                 (8,11), (11,13)]
segment_names = ["shoulder_segment",
                 "neck_segment",
                 "right_upper_arm", "right_lower_arm",
                 "left_upper_arm","left_lower_arm",
                 "pelvis_segment",
                 "trunk_segment",
                 "right_upper_leg", "right_lower_leg",
                 "left_upper_leg", "left_lower_leg"]

keypoint_names = ['right_shoulder','left_shoulder','mid_shoulder',
                  'right_wrist','left_wrist',
                  'right_elbow','left_elbow',
                  'right_hip','left_hip',
                  'mid_hip',
                  'right_knee','left_knee',
                  'right_ankle','left_ankle',
                  'head']
keypoint_number = [0,1,2,3,4,5,6,7,8,9,10,11,12,13,14]


angle_names = ["right_elbow", "left_elbow",
               "right_shoulder", "left_shoulder",
               "right_hip", "left_hip",
               "right_knee", "left_knee",
               "neck_to_shoulder", "neck_to_trunk"]

## Wrong Joint Assignment
# angle_pairs = [(3, 5, 0), (4, 6, 1),
#                (5, 0, 7), (6, 1, 8),
#                (0, 7, 9), (1, 8, 10),
#                (7, 9, 11), (8, 10, 12),
#                (13, 2, 1), (13, 2, 14)]

angle_pairs = [(3, 5, 0), (4, 6, 1),
               (5, 0, 7), (6, 1, 8),
               (0, 7, 10), (1, 8, 11),
               (7, 10, 12), (8, 11, 13),
               (14, 2, 1), (14, 2, 9)]

keypoint_dictionary = dict(zip(keypoint_names, keypoint_number))

segments_dictionary = dict(zip(segment_names, segment_pairs))

angle_dictionary = dict(zip(angle_names, angle_pairs))

master_dictionary = {}
extrinsic_angles_master_dictionary = {}

class Ui_Form(object):
    def setupUi(self, Form):
        Form.setObjectName("Form")
        Form.resize(495, 604)
        self.tabWidget = QtWidgets.QTabWidget(Form)
        self.tabWidget.setGeometry(QtCore.QRect(10, 10, 481, 585))
        self.tabWidget.setObjectName("tabWidget")
        self.tab_1 = QtWidgets.QWidget()
        self.tab_1.setObjectName("tab_1")
        self.line_5 = QtWidgets.QFrame(self.tab_1)
        self.line_5.setGeometry(QtCore.QRect(10, 160, 451, 20))
        self.line_5.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_5.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_5.setObjectName("line_5")
        self.CreateProjectLabel = QtWidgets.QLabel(self.tab_1)
        self.CreateProjectLabel.setGeometry(QtCore.QRect(10, 22, 91, 16))
        self.CreateProjectLabel.setObjectName("CreateProjectLabel")
        self.OpenPorojectPath = QtWidgets.QLineEdit(self.tab_1)
        self.OpenPorojectPath.setGeometry(QtCore.QRect(100, 18, 194, 22))
        self.OpenPorojectPath.setObjectName("OpenPorojectPath")
        self.OpenProjectButton = QtWidgets.QPushButton(self.tab_1)
        self.OpenProjectButton.setGeometry(QtCore.QRect(330, 60, 93, 28))
        self.OpenProjectButton.setObjectName("OpenProjectButton")
        self.OpenProjectButton.clicked.connect(self.Open_Project_Button_Clicked)

        self.OpenProjectBrowseButton = QtWidgets.QPushButton(self.tab_1)
        self.OpenProjectBrowseButton.setGeometry(QtCore.QRect(330, 15, 93, 28))
        self.OpenProjectBrowseButton.setObjectName("OpenProjectBrowseButton")
        self.OpenProjectBrowseButton.clicked.connect(self.Open_Project_Browse_Button_Clicked)

        self.task_name_string = QtWidgets.QLineEdit(self.tab_1)
        self.task_name_string.setGeometry(QtCore.QRect(100, 116, 194, 22))
        self.task_name_string.setObjectName("task_name_string")
        self.task_name_label = QtWidgets.QLabel(self.tab_1)
        self.task_name_label.setGeometry(QtCore.QRect(10, 120, 91, 16))
        self.task_name_label.setObjectName("task_name_label")
        self.task_create_button = QtWidgets.QPushButton(self.tab_1)
        self.task_create_button.setGeometry(QtCore.QRect(330, 114, 93, 28))
        self.task_create_button.setObjectName("task_create_button")
        self.task_create_button.clicked.connect(self.create_task_folders_clicked)

        self.Browse_Created_Project_Button = QtWidgets.QPushButton(self.tab_1)
        self.Browse_Created_Project_Button.setGeometry(QtCore.QRect(330, 195, 93, 28))
        self.Browse_Created_Project_Button.setObjectName("Browse_Created_Project_Button")
        self.Browse_Created_Project_Button.clicked.connect(self.Browse_Created_Project_Button_Clicked)

        self.Open_Created_Poroject_Path = QtWidgets.QLineEdit(self.tab_1)
        self.Open_Created_Poroject_Path.setGeometry(QtCore.QRect(100, 198, 194, 22))
        self.Open_Created_Poroject_Path.setObjectName("Open_Created_Poroject_Path")

        self.OpenProjectLabel = QtWidgets.QLabel(self.tab_1)
        self.OpenProjectLabel.setGeometry(QtCore.QRect(10, 202, 91, 16))
        self.OpenProjectLabel.setObjectName("OpenProjectLabel")
        self.Open_Created_Project_Button = QtWidgets.QPushButton(self.tab_1)
        self.Open_Created_Project_Button.setGeometry(QtCore.QRect(330, 240, 93, 28))
        self.Open_Created_Project_Button.setObjectName("Open_Created_Project_Button")
        self.Open_Created_Project_Button.clicked.connect(self.Open_Created_Project_Button_Clicked)

        self.line_6 = QtWidgets.QFrame(self.tab_1)
        self.line_6.setGeometry(QtCore.QRect(10, 280, 451, 20))
        self.line_6.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_6.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_6.setObjectName("line_6")
        self.pose_estimation_label = QtWidgets.QLabel(self.tab_1)
        self.pose_estimation_label.setGeometry(QtCore.QRect(11, 300, 171, 16))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.pose_estimation_label.setFont(font)
        self.pose_estimation_label.setObjectName("pose_estimation_label")
        self.task_video_folder_label = QtWidgets.QLabel(self.tab_1)
        self.task_video_folder_label.setGeometry(QtCore.QRect(12, 330, 91, 16))
        self.task_video_folder_label.setObjectName("task_video_folder_label")
        self.task_videos_path = QtWidgets.QLineEdit(self.tab_1)
        self.task_videos_path.setGeometry(QtCore.QRect(100, 328, 194, 22))
        self.task_videos_path.setObjectName("task_videos_path")
        self.task_videos_browse_button = QtWidgets.QPushButton(self.tab_1)
        self.task_videos_browse_button.setGeometry(QtCore.QRect(99, 360, 93, 28))
        self.task_videos_browse_button.setObjectName("task_videos_browse_button")
        self.task_videos_browse_button.clicked.connect(self.task_videos_browse_button_clicked)

        self.mediapipe_run_button = QtWidgets.QPushButton(self.tab_1)
        self.mediapipe_run_button.setGeometry(QtCore.QRect(202, 360, 93, 28))
        self.mediapipe_run_button.setObjectName("mediapipe_run_button")
        self.mediapipe_run_button.clicked.connect(self.mediapipe_run_button_clicked)

        self.skeleton_label = QtWidgets.QLabel(self.tab_1)
        self.skeleton_label.setGeometry(QtCore.QRect(310, 310, 151, 231))
        self.skeleton_label.setAlignment(QtCore.Qt.AlignCenter)
        self.skeleton_label.setText("")
        self.pixmap = QPixmap('my_skeleton.png')
        self.skeleton_label.setPixmap(
            self.pixmap.scaled(self.skeleton_label.width(), self.skeleton_label.height(),
                               Qt.KeepAspectRatio))

        self.tabWidget.addTab(self.tab_1, "")
        self.tab = QtWidgets.QWidget()
        self.tab.setObjectName("tab")
        self.task_pickle_file_label = QtWidgets.QLabel(self.tab)
        self.task_pickle_file_label.setGeometry(QtCore.QRect(10, 13, 151, 16))
        self.task_pickle_file_label.setObjectName("task_pickle_file_label")
        self.task_pickle_file_path = QtWidgets.QLineEdit(self.tab)
        self.task_pickle_file_path.setGeometry(QtCore.QRect(157, 10, 113, 22))
        self.task_pickle_file_path.setObjectName("task_pickle_file_path")
        self.task_pickle_file_browse_button = QtWidgets.QPushButton(self.tab)
        self.task_pickle_file_browse_button.setGeometry(QtCore.QRect(283, 7, 85, 28))
        self.task_pickle_file_browse_button.setObjectName("task_pickle_file_browse_button")
        self.task_pickle_file_browse_button.clicked.connect(self.task_pickle_file_browse_button_clicked)

        self.projection_matrices_pickle_file_path = QtWidgets.QLineEdit(self.tab)
        self.projection_matrices_pickle_file_path.setGeometry(QtCore.QRect(157, 53, 113, 22))
        self.projection_matrices_pickle_file_path.setObjectName("projection_matrices_pickle_file_path")
        self.projection_matrices_pickle_file_label = QtWidgets.QLabel(self.tab)
        self.projection_matrices_pickle_file_label.setGeometry(QtCore.QRect(5, 40, 151, 51))
        self.projection_matrices_pickle_file_label.setAlignment(QtCore.Qt.AlignCenter)
        self.projection_matrices_pickle_file_label.setObjectName("projection_matrices_pickle_file_label")
        self.projection_matrices_pickle_file_browse_button = QtWidgets.QPushButton(self.tab)
        self.projection_matrices_pickle_file_browse_button.setGeometry(QtCore.QRect(283, 50, 85, 28))
        self.projection_matrices_pickle_file_browse_button.setObjectName(
            "projection_matrices_pickle_file_browse_button")
        self.projection_matrices_pickle_file_browse_button.clicked.connect(self.projection_matrices_pickle_file_browse_button_clicked)

        self.reconstruct_button = QtWidgets.QPushButton(self.tab)
        self.reconstruct_button.setGeometry(QtCore.QRect(379, 7, 85, 71))
        self.reconstruct_button.setObjectName("reconstruct_button")
        self.reconstruct_button.clicked.connect(self.reconstruct_button_clicked)

        self.line = QtWidgets.QFrame(self.tab)
        self.line.setGeometry(QtCore.QRect(10, 83, 451, 16))
        self.line.setFrameShape(QtWidgets.QFrame.HLine)
        self.line.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line.setObjectName("line")
        self.data_centering_button = QtWidgets.QPushButton(self.tab)
        self.data_centering_button.setGeometry(QtCore.QRect(50, 105, 371, 28))
        self.data_centering_button.setObjectName("data_centering_button")
        self.data_centering_button.clicked.connect(self.data_centering_button_clicked)


        self.data_centering_checkBox = QtWidgets.QCheckBox(self.tab)
        self.data_centering_checkBox.setGeometry(QtCore.QRect(52, 140, 111, 20))
        self.data_centering_checkBox.setObjectName("data_centering_checkBox")


        self.data_averaging_checkBox = QtWidgets.QCheckBox(self.tab)
        self.data_averaging_checkBox.setGeometry(QtCore.QRect(52, 205, 111, 20))
        self.data_averaging_checkBox.setObjectName("data_averaging_checkBox")

        self.data_averaging_button = QtWidgets.QPushButton(self.tab)
        self.data_averaging_button.setGeometry(QtCore.QRect(50, 170, 371, 28))
        self.data_averaging_button.setObjectName("data_averaging_button")
        self.data_averaging_button.clicked.connect(self.data_averaging_button_clicked)

        self.view_keypoints_button = QtWidgets.QPushButton(self.tab)
        self.view_keypoints_button.setGeometry(QtCore.QRect(50, 238, 371, 28))
        self.view_keypoints_button.setObjectName("view_keypoints_button")
        self.view_keypoints_button.clicked.connect(self.view_keypoints_button_clicked)

        self.line_2 = QtWidgets.QFrame(self.tab)
        self.line_2.setGeometry(QtCore.QRect(10, 270, 451, 16))
        self.line_2.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_2.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_2.setObjectName("line_2")
        self.absolute_angle_calculate_button = QtWidgets.QPushButton(self.tab)
        self.absolute_angle_calculate_button.setGeometry(QtCore.QRect(50, 294, 371, 28))
        self.absolute_angle_calculate_button.setObjectName("absolute_angle_calculate_button")
        self.absolute_angle_calculate_button.clicked.connect(self.absolute_angle_calculate_button_clicked)

        self.angle_selector_comboBox = QtWidgets.QComboBox(self.tab)
        self.angle_selector_comboBox.setGeometry(QtCore.QRect(50, 333, 251, 22))
        self.angle_selector_comboBox.setObjectName("angle_selector_comboBox")
        self.angle_names = [
            "right_elbow", "left_elbow",
            "right_shoulder", "left_shoulder",
            "right_hip", "left_hip",
            "right_knee", "left_knee",
            "neck_to_shoulder", "neck_to_trunk"
        ]
        self.angle_selector_comboBox.addItems(self.angle_names)

        self.plot_absolute_angle_pushButton = QtWidgets.QPushButton(self.tab)
        self.plot_absolute_angle_pushButton.setGeometry(QtCore.QRect(320, 330, 101, 28))
        self.plot_absolute_angle_pushButton.setObjectName("plot_absolute_angle_pushButton")
        self.plot_absolute_angle_pushButton.clicked.connect(self.plot_absolute_angle_pushButton_clicked)

        self.line_3 = QtWidgets.QFrame(self.tab)
        self.line_3.setGeometry(QtCore.QRect(10, 362, 451, 16))
        self.line_3.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_3.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_3.setObjectName("line_3")
        self.segment_orientation_calculate_button = QtWidgets.QPushButton(self.tab)
        self.segment_orientation_calculate_button.setGeometry(QtCore.QRect(50, 385, 371, 28))
        self.segment_orientation_calculate_button.setObjectName("segment_orientation_calculate_button")
        self.segment_orientation_calculate_button.clicked.connect(self.segment_orientation_calculate_button_clicked)

        self.segment_orientation_plot_pushButton = QtWidgets.QPushButton(self.tab)
        self.segment_orientation_plot_pushButton.setGeometry(QtCore.QRect(320, 421, 101, 28))
        self.segment_orientation_plot_pushButton.setObjectName("segment_orientation_plot_pushButton")
        self.segment_orientation_plot_pushButton.clicked.connect(self.segment_orientation_plot_pushButton_clicked)

        self.segment_selector_comboBox = QtWidgets.QComboBox(self.tab)
        self.segment_selector_comboBox.setGeometry(QtCore.QRect(50, 424, 251, 22))
        self.segment_selector_comboBox.setObjectName("segment_selector_comboBox")
        self.segment_names_list = ["shoulder_segment",
                 "neck_segment",
                 "right_upper_arm", "right_lower_arm",
                 "left_upper_arm","left_lower_arm",
                 "pelvis_segment",
                 "trunk_segment",
                 "right_upper_leg", "right_lower_leg",
                 "left_upper_leg", "left_lower_leg"]
        self.segment_selector_comboBox.addItems(self.segment_names_list)

        self.line_4 = QtWidgets.QFrame(self.tab)
        self.line_4.setGeometry(QtCore.QRect(10, 452, 451, 16))
        self.line_4.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_4.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_4.setObjectName("line_4")
        self.animate_button = QtWidgets.QPushButton(self.tab)
        self.animate_button.setGeometry(QtCore.QRect(50, 479, 371, 28))
        self.animate_button.setObjectName("animate_button")
        self.animate_button.clicked.connect(self.animate_button_clicked)

        self.draw_coordinates_checkBox = QtWidgets.QCheckBox(self.tab)
        self.draw_coordinates_checkBox.setGeometry(QtCore.QRect(52, 520, 261, 20))
        self.draw_coordinates_checkBox.setObjectName("draw_coordinates_checkBox")

        self.tabWidget.addTab(self.tab, "")
        self.tab_2 = QtWidgets.QWidget()
        self.tab_2.setObjectName("tab_2")
        self.first_segment_selector_comboBox = QtWidgets.QComboBox(self.tab_2)
        self.first_segment_selector_comboBox.setGeometry(QtCore.QRect(110, 40, 171, 22))
        self.first_segment_selector_comboBox.setObjectName("first_segment_selector_comboBox")
        self.first_segment_selector_comboBox.addItems(self.segment_names_list)

        self.label = QtWidgets.QLabel(self.tab_2)
        self.label.setGeometry(QtCore.QRect(10, 12, 301, 16))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.first_segment_label = QtWidgets.QLabel(self.tab_2)
        self.first_segment_label.setGeometry(QtCore.QRect(9, 44, 91, 16))
        self.first_segment_label.setObjectName("first_segment_label")

        self.second_segment_selector_comboBox = QtWidgets.QComboBox(self.tab_2)
        self.second_segment_selector_comboBox.setGeometry(QtCore.QRect(110, 76, 171, 22))
        self.second_segment_selector_comboBox.setObjectName("second_segment_selector_comboBox")
        self.second_segment_selector_comboBox.addItems(self.segment_names_list)

        self.second_segment_label = QtWidgets.QLabel(self.tab_2)
        self.second_segment_label.setGeometry(QtCore.QRect(9, 80, 121, 16))
        self.second_segment_label.setObjectName("second_segment_label")

        self.intersegment_calculate_pushButton = QtWidgets.QPushButton(self.tab_2)
        self.intersegment_calculate_pushButton.setGeometry(QtCore.QRect(370, 37, 93, 28))
        self.intersegment_calculate_pushButton.setObjectName("intersegment_calculate_pushButton")
        self.intersegment_calculate_pushButton.clicked.connect(self.intersegment_calculate_pushButton_clicked)

        self.intersegment_angle_save_pushButton = QtWidgets.QPushButton(self.tab_2)
        self.intersegment_angle_save_pushButton.setGeometry(QtCore.QRect(370, 74, 93, 28))
        self.intersegment_angle_save_pushButton.setObjectName("intersegment_angle_save_pushButton")
        self.intersegment_angle_save_pushButton.clicked.connect(self.intersegment_angle_save_pushButton_clicked)

        self.line_7 = QtWidgets.QFrame(self.tab_2)
        self.line_7.setGeometry(QtCore.QRect(10, 102, 451, 16))
        self.line_7.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_7.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_7.setObjectName("line_7")
        self.label_2 = QtWidgets.QLabel(self.tab_2)
        self.label_2.setGeometry(QtCore.QRect(10, 121, 421, 16))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.planar_angle_values_plot_pushButton = QtWidgets.QPushButton(self.tab_2)
        self.planar_angle_values_plot_pushButton.setGeometry(QtCore.QRect(310, 186, 101, 28))
        self.planar_angle_values_plot_pushButton.setObjectName("planar_angle_values_plot_pushButton")
        self.planar_angle_values_plot_pushButton.clicked.connect(self.planar_angle_values_plot_pushButton_clicked)

        self.planar_angle_selector_comboBox = QtWidgets.QComboBox(self.tab_2)
        self.planar_angle_selector_comboBox.setGeometry(QtCore.QRect(40, 189, 251, 22))
        self.planar_angle_selector_comboBox.setObjectName("planar_angle_selector_comboBox")
        self.planar_angle_selector_comboBox.addItems(self.angle_names)

        self.planar_angle_values_calculate_button = QtWidgets.QPushButton(self.tab_2)
        self.planar_angle_values_calculate_button.setGeometry(QtCore.QRect(40, 150, 371, 28))
        self.planar_angle_values_calculate_button.setObjectName("planar_angle_values_calculate_button")
        self.planar_angle_values_calculate_button.clicked.connect(self.planar_angle_values_calculate_button_clicked)

        self.line_8 = QtWidgets.QFrame(self.tab_2)
        self.line_8.setGeometry(QtCore.QRect(10, 216, 451, 16))
        self.line_8.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_8.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_8.setObjectName("line_8")
        self.label_3 = QtWidgets.QLabel(self.tab_2)
        self.label_3.setGeometry(QtCore.QRect(10, 232, 401, 16))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")
        self.euler_angle_save_pushButton = QtWidgets.QPushButton(self.tab_2)
        self.euler_angle_save_pushButton.setGeometry(QtCore.QRect(350, 294, 93, 28))
        self.euler_angle_save_pushButton.setObjectName("euler_angle_save_pushButton")
        self.euler_angle_save_pushButton.clicked.connect(self.euler_angle_save_pushButton_clicked)

        self.first_segment_euler_label = QtWidgets.QLabel(self.tab_2)
        self.first_segment_euler_label.setGeometry(QtCore.QRect(9, 264, 91, 16))
        self.first_segment_euler_label.setObjectName("first_segment_euler_label")
        self.second_segment_euler_label = QtWidgets.QLabel(self.tab_2)
        self.second_segment_euler_label.setGeometry(QtCore.QRect(9, 300, 121, 16))
        self.second_segment_euler_label.setObjectName("second_segment_euler_label")
        self.euler_calculate_pushButton = QtWidgets.QPushButton(self.tab_2)
        self.euler_calculate_pushButton.setGeometry(QtCore.QRect(350, 257, 93, 28))
        self.euler_calculate_pushButton.setObjectName("euler_calculate_pushButton")
        self.euler_calculate_pushButton.clicked.connect(self.euler_calculate_pushButton_clicked)

        self.first_segment_selector_euler_comboBox = QtWidgets.QComboBox(self.tab_2)
        self.first_segment_selector_euler_comboBox.setGeometry(QtCore.QRect(120, 260, 211, 22))
        self.first_segment_selector_euler_comboBox.setObjectName("first_segment_selector_euler_comboBox")
        self.first_segment_selector_euler_comboBox.addItems(self.segment_names_list)

        self.second_segment_selector_euler_comboBox = QtWidgets.QComboBox(self.tab_2)
        self.second_segment_selector_euler_comboBox.setGeometry(QtCore.QRect(120, 296, 211, 22))
        self.second_segment_selector_euler_comboBox.setObjectName("second_segment_selector_euler_comboBox")
        self.second_segment_selector_euler_comboBox.addItems(self.segment_names_list)

        self.line_9 = QtWidgets.QFrame(self.tab_2)
        self.line_9.setGeometry(QtCore.QRect(10, 326, 451, 16))
        self.line_9.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_9.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_9.setObjectName("line_9")

        self.first_segment_reverse_checkBox = QtWidgets.QCheckBox(self.tab_2)
        self.first_segment_reverse_checkBox.setGeometry(QtCore.QRect(290, 40, 81, 20))
        self.first_segment_reverse_checkBox.setObjectName("first_segment_reverse_checkBox")
        self.second_segment_reverse_checkBox = QtWidgets.QCheckBox(self.tab_2)
        self.second_segment_reverse_checkBox.setGeometry(QtCore.QRect(290, 77, 81, 20))
        self.second_segment_reverse_checkBox.setObjectName("second_segment_reverse_checkBox")

        self.tabWidget.addTab(self.tab_2, "")

        self.retranslateUi(Form)
        self.tabWidget.setCurrentIndex(0)
        QtCore.QMetaObject.connectSlotsByName(Form)

    def retranslateUi(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "PoseTrack"))
        self.CreateProjectLabel.setText(_translate("Form", "Create Project"))
        self.OpenProjectButton.setText(_translate("Form", "Open"))
        self.OpenProjectBrowseButton.setText(_translate("Form", "Browse"))
        self.task_name_label.setText(_translate("Form", "Task Name"))
        self.task_create_button.setText(_translate("Form", "Create"))
        self.Browse_Created_Project_Button.setText(_translate("Form", "Browse"))
        self.OpenProjectLabel.setText(_translate("Form", "Open Project"))
        self.Open_Created_Project_Button.setText(_translate("Form", "Open"))
        self.pose_estimation_label.setText(_translate("Form", "MediaPipe Pose Estimation "))
        self.task_video_folder_label.setText(_translate("Form", "Videos Folder"))
        self.task_videos_browse_button.setText(_translate("Form", "Browse"))
        self.mediapipe_run_button.setText(_translate("Form", "Run"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_1), _translate("Form", "Project"))
        self.task_pickle_file_label.setText(_translate("Form", "Load Tasks Pickle Files"))
        self.task_pickle_file_browse_button.setText(_translate("Form", "Browse"))
        self.projection_matrices_pickle_file_label.setText(_translate("Form", "Load Projection Matrices\n"
                                                                              "Pickle Files"))
        self.projection_matrices_pickle_file_browse_button.setText(_translate("Form", "Browse"))
        self.reconstruct_button.setText(_translate("Form", "Reconstruct"))
        self.data_centering_button.setText(_translate("Form", "Data Centering and Outlier Removal Process"))
        self.data_centering_checkBox.setText(_translate("Form", "View Results"))
        self.data_averaging_checkBox.setText(_translate("Form", "View Results"))
        self.data_averaging_button.setText(_translate("Form", "Data Enhancement"))
        self.view_keypoints_button.setText(_translate("Form", "View Keypoints Coordinates"))
        self.absolute_angle_calculate_button.setText(_translate("Form", "Calculate Absolute Joint Angles"))
        self.plot_absolute_angle_pushButton.setText(_translate("Form", "Plot"))
        self.segment_orientation_calculate_button.setText(
            _translate("Form", "Calculate Extrinsic Segment Orientations"))
        self.segment_orientation_plot_pushButton.setText(_translate("Form", "Plot"))
        self.animate_button.setText(_translate("Form", "Create Animate Model"))
        self.draw_coordinates_checkBox.setText(_translate("Form", "Generate Segment Coordinate Frame"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab), _translate("Form", "Data Processing"))
        self.label.setText(_translate("Form", "Calculate InterSegment (Intrinsic) Angles"))
        self.first_segment_label.setText(_translate("Form", "First Segment"))
        self.second_segment_label.setText(_translate("Form", "Second Segment"))
        self.intersegment_calculate_pushButton.setText(_translate("Form", "Calculate"))
        self.intersegment_angle_save_pushButton.setText(_translate("Form", "Save"))
        self.label_2.setText(_translate("Form", "Planar Values Of Joint Angles in Global Coordinate Frame"))
        self.planar_angle_values_plot_pushButton.setText(_translate("Form", "Plot"))
        self.planar_angle_values_calculate_button.setText(
            _translate("Form", "Calculate Planar Values of Joints Angles"))
        self.label_3.setText(_translate("Form", "Calculate Relative Euler Angles"))
        self.euler_angle_save_pushButton.setText(_translate("Form", "Save"))
        self.first_segment_euler_label.setText(_translate("Form", "First Segment"))
        self.second_segment_euler_label.setText(_translate("Form", "Second Segment"))
        self.euler_calculate_pushButton.setText(_translate("Form", "Calculate"))
        self.first_segment_reverse_checkBox.setText(_translate("Form", "Reverse"))
        self.second_segment_reverse_checkBox.setText(_translate("Form", "Reverse"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_2), _translate("Form", "Data Processing (Continued)"))

    def Open_Project_Browse_Button_Clicked(self):
        global ProjectOpenPath
        self.dirname = QFileDialog.getExistingDirectory(None, "Select Folder", "./")
        self.OpenPorojectPath.setText(self.dirname)
        ProjectOpenPath = self.dirname

    def Open_Project_Button_Clicked(self):
        global ProjectWorkingPath

        ProjectWorkingPath = ProjectOpenPath
        print(ProjectWorkingPath)

    def create_task_folders_clicked(self):
        global ProjectWorkingPath

        def create_dir(path):
            if not os.path.exists(path):
                os.makedirs(path)
            return path  # Return the created directory path

        save_dir = os.path.join(ProjectWorkingPath, "Motion_Analysis")
        task_name = self.task_name_string.text()
        if task_name =="":
            print("Enter Task Name")
        else:
            save_dir = create_dir(os.path.join(save_dir, task_name))

            for item in ["task_videos", "task_video_results", "task_pickle_files", "task_result_files"]:
                create_dir(os.path.join(save_dir, item))

            ProjectWorkingPath = save_dir
            print(ProjectWorkingPath)

    def Browse_Created_Project_Button_Clicked(self):
        global Open_Created_Poroject_Path
        self.dirname = QFileDialog.getExistingDirectory(None, "Select Folder", "./")
        self.Open_Created_Poroject_Path.setText(self.dirname)
        Open_Created_Poroject_Path = self.dirname

    def Open_Created_Project_Button_Clicked(self):
        global ProjectWorkingPath

        ProjectWorkingPath = Open_Created_Poroject_Path
        print(ProjectWorkingPath)

    def task_videos_browse_button_clicked(self):
        global task_videos_path_for_pose_estimation
        self.dirname = QFileDialog.getExistingDirectory(None, "Select Folder", ProjectWorkingPath + "/task_videos")
        self.task_videos_path.setText(self.dirname)
        task_videos_path_for_pose_estimation = self.dirname
        print(task_videos_path_for_pose_estimation)

    def mediapipe_run_button_clicked(self):
        # Initialize MediaPipe Pose
        mp_pose = mp.solutions.pose
        pose = mp_pose.Pose()
        # Iterate through the videos in the folder
        for video_file in os.listdir(task_videos_path_for_pose_estimation):
            if video_file.endswith('.mp4'):
                video_path = os.path.join(task_videos_path_for_pose_estimation, video_file)

                cap = cv2.VideoCapture(video_path)

                # Extract video file name and append "_mpOut"
                input_file_name = os.path.basename(video_path)
                file_name_without_ext = os.path.splitext(input_file_name)[0]
                output_file_name = file_name_without_ext + "_mpOut.mp4"

                # Get the original video's width and height
                original_width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
                original_height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))

                # Calculate the aspect ratio
                aspect_ratio = original_width / original_height

                # Set the height for the resized video
                resized_height = 720
                resized_width = int(resized_height * aspect_ratio)

                # Define the codec and create a VideoWriter object
                fourcc = cv2.VideoWriter_fourcc(*'mp4v')  # or use 'XVID'
                output_file_path = os.path.join(ProjectWorkingPath, "task_video_results", output_file_name)
                out = cv2.VideoWriter(output_file_path, fourcc, 30.0, (resized_width, resized_height))

                # Dictionary to store keypoints coordinates
                keypoints_data = {
                    'right_shoulder': [],
                    'left_shoulder': [],
                    'mid_shoulder': [],
                    'right_wrist': [],
                    'left_wrist': [],
                    'right_elbow': [],
                    'left_elbow': [],
                    'right_hip': [],
                    'left_hip': [],
                    'mid_hip': [],
                    'right_knee': [],
                    'left_knee': [],
                    'right_ankle': [],
                    'left_ankle': [],
                    'head': []
                }

                frame_number = 0  # Initialize the frame number

                while cap.isOpened():
                    ret, frame = cap.read()
                    if not ret:
                        break

                    # Resize the frame
                    frame = cv2.resize(frame, (resized_width, resized_height), interpolation=cv2.INTER_AREA)

                    # Convert the BGR image to RGB
                    image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

                    start_time = time.time()

                    # Process the image and find the pose landmarks
                    results = pose.process(image)

                    # Draw the pose landmarks and segments on the image
                    if results.pose_landmarks:
                        # Draw selected pose landmarks
                        landmarks_to_draw = [mp_pose.PoseLandmark.LEFT_WRIST,
                                             mp_pose.PoseLandmark.LEFT_ELBOW,
                                             mp_pose.PoseLandmark.LEFT_SHOULDER,
                                             mp_pose.PoseLandmark.RIGHT_SHOULDER,
                                             mp_pose.PoseLandmark.RIGHT_ELBOW,
                                             mp_pose.PoseLandmark.RIGHT_WRIST,
                                             mp_pose.PoseLandmark.LEFT_HIP,
                                             mp_pose.PoseLandmark.RIGHT_HIP,
                                             mp_pose.PoseLandmark.LEFT_KNEE,
                                             mp_pose.PoseLandmark.RIGHT_KNEE]

                        mid_shoulder = [(results.pose_landmarks.landmark[mp_pose.PoseLandmark.LEFT_SHOULDER].x +
                                         results.pose_landmarks.landmark[mp_pose.PoseLandmark.RIGHT_SHOULDER].x) / 2,
                                        (results.pose_landmarks.landmark[mp_pose.PoseLandmark.LEFT_SHOULDER].y +
                                         results.pose_landmarks.landmark[mp_pose.PoseLandmark.RIGHT_SHOULDER].y) / 2]

                        mid_hip = [(results.pose_landmarks.landmark[mp_pose.PoseLandmark.LEFT_HIP].x +
                                    results.pose_landmarks.landmark[mp_pose.PoseLandmark.RIGHT_HIP].x) / 2,
                                   (results.pose_landmarks.landmark[mp_pose.PoseLandmark.LEFT_HIP].y +
                                    results.pose_landmarks.landmark[mp_pose.PoseLandmark.RIGHT_HIP].y) / 2]

                        head = [(results.pose_landmarks.landmark[mp_pose.PoseLandmark.LEFT_EAR].x +
                                 results.pose_landmarks.landmark[mp_pose.PoseLandmark.RIGHT_EAR].x) / 2,
                                (results.pose_landmarks.landmark[mp_pose.PoseLandmark.LEFT_EAR].y +
                                 results.pose_landmarks.landmark[mp_pose.PoseLandmark.RIGHT_EAR].y) / 2]

                        keypoints_data['right_shoulder'].append({'frame': frame_number, 'coordinates':
                            [int(results.pose_landmarks.landmark[mp_pose.PoseLandmark.RIGHT_SHOULDER].x * frame.shape[
                                1]),
                             int(results.pose_landmarks.landmark[mp_pose.PoseLandmark.RIGHT_SHOULDER].y * frame.shape[
                                 0])]})

                        keypoints_data['left_shoulder'].append({'frame': frame_number, 'coordinates':
                            [int(
                                results.pose_landmarks.landmark[mp_pose.PoseLandmark.LEFT_SHOULDER].x * frame.shape[1]),
                             int(results.pose_landmarks.landmark[mp_pose.PoseLandmark.LEFT_SHOULDER].y * frame.shape[
                                 0])]})

                        keypoints_data['mid_shoulder'].append({'frame': frame_number, 'coordinates':
                            [int(mid_shoulder[0] * frame.shape[1]),
                             int(mid_shoulder[1] * frame.shape[0])]})

                        keypoints_data['head'].append({'frame': frame_number, 'coordinates':
                            [int(head[0] * frame.shape[1]),
                             int(head[1] * frame.shape[0])]})

                        keypoints_data['right_wrist'].append({'frame': frame_number, 'coordinates':
                            [int(results.pose_landmarks.landmark[mp_pose.PoseLandmark.RIGHT_WRIST].x * frame.shape[1]),
                             int(results.pose_landmarks.landmark[mp_pose.PoseLandmark.RIGHT_WRIST].y * frame.shape[
                                 0])]})

                        keypoints_data['left_wrist'].append({'frame': frame_number, 'coordinates':
                            [int(results.pose_landmarks.landmark[mp_pose.PoseLandmark.LEFT_WRIST].x * frame.shape[1]),
                             int(results.pose_landmarks.landmark[mp_pose.PoseLandmark.LEFT_WRIST].y * frame.shape[0])]})

                        keypoints_data['right_elbow'].append({'frame': frame_number, 'coordinates':
                            [int(results.pose_landmarks.landmark[mp_pose.PoseLandmark.RIGHT_ELBOW].x * frame.shape[1]),
                             int(results.pose_landmarks.landmark[mp_pose.PoseLandmark.RIGHT_ELBOW].y * frame.shape[
                                 0])]})

                        keypoints_data['left_elbow'].append({'frame': frame_number, 'coordinates':
                            [int(results.pose_landmarks.landmark[mp_pose.PoseLandmark.LEFT_ELBOW].x * frame.shape[1]),
                             int(results.pose_landmarks.landmark[mp_pose.PoseLandmark.LEFT_ELBOW].y * frame.shape[0])]})

                        keypoints_data['right_hip'].append({'frame': frame_number, 'coordinates':
                            [int(results.pose_landmarks.landmark[mp_pose.PoseLandmark.RIGHT_HIP].x * frame.shape[1]),
                             int(results.pose_landmarks.landmark[mp_pose.PoseLandmark.RIGHT_HIP].y * frame.shape[0])]})

                        keypoints_data['left_hip'].append({'frame': frame_number, 'coordinates':
                            [int(results.pose_landmarks.landmark[mp_pose.PoseLandmark.LEFT_HIP].x * frame.shape[1]),
                             int(results.pose_landmarks.landmark[mp_pose.PoseLandmark.LEFT_HIP].y * frame.shape[0])]})

                        keypoints_data['mid_hip'].append({'frame': frame_number, 'coordinates':
                            [int(mid_hip[0] * frame.shape[1]),
                             int(mid_hip[1] * frame.shape[0])]})

                        keypoints_data['right_knee'].append({'frame': frame_number, 'coordinates':
                            [int(results.pose_landmarks.landmark[mp_pose.PoseLandmark.RIGHT_KNEE].x * frame.shape[1]),
                             int(results.pose_landmarks.landmark[mp_pose.PoseLandmark.RIGHT_KNEE].y * frame.shape[0])]})

                        keypoints_data['left_knee'].append({'frame': frame_number, 'coordinates':
                            [int(results.pose_landmarks.landmark[mp_pose.PoseLandmark.LEFT_KNEE].x * frame.shape[1]),
                             int(results.pose_landmarks.landmark[mp_pose.PoseLandmark.LEFT_KNEE].y * frame.shape[0])]})

                        keypoints_data['right_ankle'].append({'frame': frame_number, 'coordinates':
                            [int(results.pose_landmarks.landmark[mp_pose.PoseLandmark.RIGHT_ANKLE].x * frame.shape[1]),
                             int(results.pose_landmarks.landmark[mp_pose.PoseLandmark.RIGHT_ANKLE].y * frame.shape[
                                 0])]})

                        keypoints_data['left_ankle'].append({'frame': frame_number, 'coordinates':
                            [int(results.pose_landmarks.landmark[mp_pose.PoseLandmark.LEFT_ANKLE].x * frame.shape[1]),
                             int(results.pose_landmarks.landmark[mp_pose.PoseLandmark.LEFT_ANKLE].y * frame.shape[0])]})

                        for landmark in landmarks_to_draw:
                            x, y = int(results.pose_landmarks.landmark[landmark].x * frame.shape[1]), int(
                                results.pose_landmarks.landmark[landmark].y * frame.shape[0])
                            cv2.circle(frame, (x, y), 3, (0, 0, 255), -1)

                        # Draw segments between the selected pose landmarks
                        connections = [[mp_pose.PoseLandmark.LEFT_WRIST, mp_pose.PoseLandmark.LEFT_ELBOW],
                                       [mp_pose.PoseLandmark.LEFT_ELBOW, mp_pose.PoseLandmark.LEFT_SHOULDER],
                                       [mp_pose.PoseLandmark.LEFT_SHOULDER, mp_pose.PoseLandmark.RIGHT_SHOULDER],
                                       [mp_pose.PoseLandmark.RIGHT_SHOULDER, mp_pose.PoseLandmark.RIGHT_ELBOW],
                                       [mp_pose.PoseLandmark.RIGHT_ELBOW, mp_pose.PoseLandmark.RIGHT_WRIST],
                                       # [mp_pose.PoseLandmark.LEFT_SHOULDER, mp_pose.PoseLandmark.LEFT_HIP],
                                       # [mp_pose.PoseLandmark.RIGHT_SHOULDER, mp_pose.PoseLandmark.RIGHT_HIP],
                                       [mp_pose.PoseLandmark.LEFT_HIP, mp_pose.PoseLandmark.RIGHT_HIP],
                                       [mp_pose.PoseLandmark.LEFT_HIP, mp_pose.PoseLandmark.LEFT_KNEE],
                                       [mp_pose.PoseLandmark.RIGHT_HIP, mp_pose.PoseLandmark.RIGHT_KNEE],
                                       [mp_pose.PoseLandmark.LEFT_KNEE, mp_pose.PoseLandmark.LEFT_ANKLE],
                                       [mp_pose.PoseLandmark.RIGHT_KNEE, mp_pose.PoseLandmark.RIGHT_ANKLE]]

                        for connection in connections:
                            start_x, start_y = int(
                                results.pose_landmarks.landmark[connection[0]].x * frame.shape[1]), int(
                                results.pose_landmarks.landmark[connection[0]].y * frame.shape[0])
                            end_x, end_y = int(results.pose_landmarks.landmark[connection[1]].x * frame.shape[1]), int(
                                results.pose_landmarks.landmark[connection[1]].y * frame.shape[0])
                            cv2.line(frame, (start_x, start_y), (end_x, end_y), (255, 255, 255), 2)

                            # Now, let's draw the "mid_shoulder" connection
                            mid_shoulder = [(results.pose_landmarks.landmark[mp_pose.PoseLandmark.LEFT_SHOULDER].x +
                                             results.pose_landmarks.landmark[
                                                 mp_pose.PoseLandmark.RIGHT_SHOULDER].x) / 2,
                                            (results.pose_landmarks.landmark[mp_pose.PoseLandmark.LEFT_SHOULDER].y +
                                             results.pose_landmarks.landmark[
                                                 mp_pose.PoseLandmark.RIGHT_SHOULDER].y) / 2]

                            mid_hip = [(results.pose_landmarks.landmark[mp_pose.PoseLandmark.LEFT_HIP].x +
                                        results.pose_landmarks.landmark[mp_pose.PoseLandmark.RIGHT_HIP].x) / 2,
                                       (results.pose_landmarks.landmark[mp_pose.PoseLandmark.LEFT_HIP].y +
                                        results.pose_landmarks.landmark[mp_pose.PoseLandmark.RIGHT_HIP].y) / 2]

                            head = [(results.pose_landmarks.landmark[mp_pose.PoseLandmark.LEFT_EAR].x +
                                     results.pose_landmarks.landmark[mp_pose.PoseLandmark.RIGHT_EAR].x) / 2,
                                    (results.pose_landmarks.landmark[mp_pose.PoseLandmark.LEFT_EAR].y +
                                     results.pose_landmarks.landmark[mp_pose.PoseLandmark.RIGHT_EAR].y) / 2]

                            # Draw mid_shoulder and mid_shoulder to mid_hip lines in white
                            cv2.line(frame,
                                     (int(mid_shoulder[0] * frame.shape[1]), int(mid_shoulder[1] * frame.shape[0])),
                                     (int(head[0] * frame.shape[1]), int(head[1] * frame.shape[0])),
                                     (255, 255, 255), 2)

                            cv2.line(frame,
                                     (int(mid_shoulder[0] * frame.shape[1]), int(mid_shoulder[1] * frame.shape[0])),
                                     (int(mid_hip[0] * frame.shape[1]), int(mid_hip[1] * frame.shape[0])),
                                     (255, 255, 255), 2)

                            # Draw joints (shoulder, mid_shoulder, and middle of ears) in red circles
                            cv2.circle(frame,
                                       (int(mid_shoulder[0] * frame.shape[1]), int(mid_shoulder[1] * frame.shape[0])),
                                       3, (0, 0, 255), -1)
                            cv2.circle(frame, (int(mid_hip[0] * frame.shape[1]), int(mid_hip[1] * frame.shape[0])), 3,
                                       (0, 0, 255), -1)
                            cv2.circle(frame, (int(head[0] * frame.shape[1]), int(head[1] * frame.shape[0])), 3,
                                       (0, 0, 255), -1)

                        frame_number += 1
                    end_time = time.time()
                    fps = 1 / (end_time - start_time)
                    print("FPS :", fps)

                    cv2.putText(frame, "FPS :" + str(int(fps)), (10, 50), cv2.FONT_HERSHEY_COMPLEX, 1.2, (255, 0, 255),
                                1,
                                cv2.LINE_AA)

                    # Write the frame into the file 'output.mp4'
                    out.write(frame)

                    # Display the annotated frame
                    cv2.imshow('MediaPipe Pose', frame)

                    # Break the loop if 'q' is pressed
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break

                cap.release()
                out.release()  # Save the output
                cv2.destroyAllWindows()

                # Save keypoints_data in a pickle file with the name of the input video
                pickle_file_name = file_name_without_ext + "_keypoints_data.pkl"
                pickle_file_path = os.path.join(ProjectWorkingPath, "task_pickle_files", pickle_file_name)
                with open(pickle_file_path, 'wb') as f:
                    pickle.dump(keypoints_data, f)

                print(f"Keypoints data saved in pickle file: {pickle_file_path}")

    def task_pickle_file_browse_button_clicked(self):
        global tasks_pickle_files_directory
        global task_file_path_list
        self.dirname = QFileDialog.getExistingDirectory(None, "Select Folder", ProjectWorkingPath + "/task_pickle_files")
        self.task_pickle_file_path.setText(self.dirname)
        tasks_pickle_files_directory = self.dirname
        print(tasks_pickle_files_directory)
        task_file_path_list = list_files_without_prefix(tasks_pickle_files_directory)
        print(task_file_path_list)

    def projection_matrices_pickle_file_browse_button_clicked(self):
        global projection_matrices_directory
        global projection_file_list
        parent_directory = os.path.abspath(os.path.join(ProjectWorkingPath, os.pardir, os.pardir))

        self.dirname = QFileDialog.getExistingDirectory(None, "Select Folder",
                                                        parent_directory)
        self.projection_matrices_pickle_file_path.setText(self.dirname)
        projection_matrices_directory = self.dirname
        print(projection_matrices_directory)
        projection_file_list = list_files_with_prefix(projection_matrices_directory)
        print(projection_file_list)

    def reconstruct_button_clicked(self):
        global transformed_dictionary
        transformed_dictionary = three_dimensional_reconstruction(task_file_path_list, projection_file_list)

    def data_centering_button_clicked(self):
        global centralized_series
        print("Processing is started!!!")
        if self.data_centering_checkBox.isChecked():
            centralized_series = data_centering_and_outlier_removal(transformed_dictionary)
            plot_centralized_series(centralized_series, keypoint_names)
        else:
            centralized_series = data_centering_and_outlier_removal(transformed_dictionary)

    def data_averaging_button_clicked(self):
        print("Averaging is started!!!")
        if self.data_averaging_checkBox.isChecked():
            average_data_dict = plot_centralized_series_with_average(centralized_series, keypoint_names)
        else:
            average_data_dict = data_averaging_procedure(centralized_series)

        # Save keypoints_data in a pickle file named "enhanced_keypoints_coordinates.pkl"
        with open(ProjectWorkingPath + '/task_result_files/enhanced_keypoints_coordinates.pkl', "wb") as f:
            pickle.dump(average_data_dict, f)

        print("Enhanced keypoints coordinates saved in : enhanced_keypoints_coordinates.pkl")

    def view_keypoints_button_clicked(self):
        with open(ProjectWorkingPath + '/task_result_files/enhanced_keypoints_coordinates.pkl', "rb") as f:
            average_data_dict = pickle.load(f)

        for key, value in average_data_dict.items():
            # Create a new figure for each key
            plt.figure(figsize=(12, 10))
            plt.suptitle(keypoint_names[key])  # Set the figure title

            for dim in range(value.shape[1]):
                plt.subplot(3, 1, dim + 1)
                # Plot the averaged data with a solid black line
                plt.plot(value[:, dim], color='black')

            plt.tight_layout()
            plt.show()

    def absolute_angle_calculate_button_clicked(self):
        global absolute_joint_angles
        with open(ProjectWorkingPath + '/task_result_files/enhanced_keypoints_coordinates.pkl', "rb") as f:
            average_data_dict = pickle.load(f)

        keypoints_coordinates_data = []
        for value in average_data_dict.values():
            keypoints_coordinates_data.append(value)

        keypoints_coordinates_data = np.array(keypoints_coordinates_data)
        # Now keypoints_3D_filtered has the shape (number_of_segments, number_of_video_frames, 3)
        # print(keypoints_coordinates_data.shape)

        absolute_joint_angles = calculate_absolute_angles(keypoints_coordinates_data, angle_dictionary)
        # Save absolute_joint_angles in a pickle file named "absolute_joint_angles.pkl"
        with open(ProjectWorkingPath + '/task_result_files/absolute_joint_angles.pkl', "wb") as f:
            pickle.dump(absolute_joint_angles, f)

        print("Absolute joints angles were calculated.")

    def plot_absolute_angle_pushButton_clicked(self):
        selected_angle = self.angle_selector_comboBox.currentText()

        if selected_angle in absolute_joint_angles[0].keys():
            print(f"Plotting {selected_angle}")
            plt.figure()  # Create a new figure
            plt.title(selected_angle)  # Set the title to the selected angle
            plt.plot(absolute_joint_angles[0][selected_angle])
            plt.show()

    def segment_orientation_calculate_button_clicked(self):
        global extrinsic_segment_orientations
        global formatted_extrinsic_segment_orientations_to_csv
        with open(ProjectWorkingPath + '/task_result_files/enhanced_keypoints_coordinates.pkl', "rb") as f:
            average_data_dict = pickle.load(f)

        keypoints_coordinates_data = []
        for value in average_data_dict.values():
            keypoints_coordinates_data.append(value)

        keypoints_coordinates_data = np.array(keypoints_coordinates_data)
        # Now keypoints_3D_filtered has the shape (number_of_segments, number_of_video_frames, 3)
        # print(keypoints_coordinates_data.shape)

        extrinsic_segment_orientations = calculate_extrinsic_segment_angles(keypoints_coordinates_data, segments_dictionary)

        formatted_extrinsic_segment_orientations_to_csv = {}

        for frame, frame_data in extrinsic_segment_orientations.items():
            for segment, angles in frame_data.items():
                if segment not in formatted_extrinsic_segment_orientations_to_csv:
                    formatted_extrinsic_segment_orientations_to_csv[segment] = {}
                formatted_extrinsic_segment_orientations_to_csv[segment][frame] = angles

        # Define the file path
        excel_file_path = os.path.join(ProjectWorkingPath, "task_result_files", 'global_segment_orientation.xlsx')
        pickle_file_path = os.path.join(ProjectWorkingPath, "task_result_files", 'extrinsic_segment_orientations.pkl')

        # Check if the file already exists
        if not os.path.exists(excel_file_path):
            # Save extrinsic angles to Excel
            save_extrinsic_angles_to_excel(formatted_extrinsic_segment_orientations_to_csv, excel_file_path)

            with open(pickle_file_path, 'wb') as f:
                pickle.dump(extrinsic_segment_orientations, f)

            print("Extrinsic segments orientations (angles) were calculated and saved to Excel.")
        else:
            print("The Excel file already exists. Skipping save operation.")

    def segment_orientation_plot_pushButton_clicked(self):
        selected_segment = self.segment_selector_comboBox.currentText()

        if selected_segment in formatted_extrinsic_segment_orientations_to_csv.keys():
            print(f"Plotting {selected_segment}")
            plot_segment_angles(selected_segment, formatted_extrinsic_segment_orientations_to_csv[selected_segment])

    def animate_button_clicked(self):

        with open(ProjectWorkingPath + '/task_result_files/enhanced_keypoints_coordinates.pkl', "rb") as f:
            average_data_dict = pickle.load(f)

        keypoints_coordinates_data = []
        for value in average_data_dict.values():
            keypoints_coordinates_data.append(value)

        keypoints_coordinates_data = np.array(keypoints_coordinates_data)
        print(keypoints_coordinates_data.shape)

        if self.draw_coordinates_checkBox.isChecked():
            animate_body_skeleton_with_coordinate_frame(keypoints_coordinates_data, segment_pairs)
        else:
            with open(ProjectWorkingPath + '/task_result_files/absolute_joint_angles.pkl', "rb") as f:
                absolute_joint_angles = pickle.load(f)
                joint_angles_frame_by_frame = absolute_joint_angles[1]
            animate_body_skeleton_wo_coordinate_frame(keypoints_coordinates_data, segment_pairs, joint_angles_frame_by_frame)

    def intersegment_calculate_pushButton_clicked(self):
        global desired_angle_rotation_differences
        first_selected_segment = self.first_segment_selector_comboBox.currentText()
        second_selected_segment = self.second_segment_selector_comboBox.currentText()

        if self.first_segment_reverse_checkBox.isChecked():
            first_scale = -1
        else:
            first_scale = 1

        if self.second_segment_reverse_checkBox.isChecked():
            second_scale = -1
        else:
            second_scale = 1

        with open(ProjectWorkingPath + '/task_result_files/extrinsic_segment_orientations.pkl', "rb") as f:
            all_extrinsic_angles = pickle.load(f)

        desired_angle_rotation_differences = calculate_intesegment_angles(first_selected_segment, second_selected_segment, all_extrinsic_angles, first_scale, second_scale)

        self.first_segment_reverse_checkBox.setChecked(False)
        self.second_segment_reverse_checkBox.setChecked(False)

    def intersegment_angle_save_pushButton_clicked(self):
        first_selected_segment = self.first_segment_selector_comboBox.currentText()
        second_selected_segment = self.second_segment_selector_comboBox.currentText()

        # Define the file path
        excel_file_path = os.path.join(ProjectWorkingPath, "task_result_files", f'{first_selected_segment}_{second_selected_segment}.xlsx')
        pickle_file_path = os.path.join(ProjectWorkingPath, "task_result_files", f'{first_selected_segment}_{second_selected_segment}.pkl')

        # Check if the file already exists
        if not os.path.exists(excel_file_path):
            # Save extrinsic angles to Excel
            save_intrinsic_angles_to_excel(desired_angle_rotation_differences, excel_file_path)

            with open(pickle_file_path, 'wb') as f:
                pickle.dump(desired_angle_rotation_differences, f)

            print(f"intersegment angles between {first_selected_segment} and {second_selected_segment} were calculated and saved to Excel.")
        else:
            print("The Excel file already exists. Skipping save operation.")

    def planar_angle_values_calculate_button_clicked(self):

        with open(ProjectWorkingPath + '/task_result_files/enhanced_keypoints_coordinates.pkl', "rb") as f:
            average_data_dict = pickle.load(f)

        keypoints_coordinates_data = []
        for value in average_data_dict.values():
            keypoints_coordinates_data.append(value)

        keypoints_coordinates_data = np.array(keypoints_coordinates_data)

        transformed_all_planar_joints_angles = calculate_planar_joints_angles(keypoints_coordinates_data,
                                                                              angle_dictionary)

        planar_joint_angles = {}

        for key, values in transformed_all_planar_joints_angles.items():
            x_rotation = []
            y_rotation = []
            z_rotation = []

            for value in values:
                x_rotation.append(value[0])
                y_rotation.append(value[1])
                z_rotation.append(value[2])

            planar_joint_angles[key] = np.array([x_rotation, y_rotation, z_rotation]).T

        # Save keypoints_data in a pickle file named "enhanced_keypoints_coordinates.pkl"
        with open(ProjectWorkingPath + '/task_result_files/planar_global_joints_angles.pkl', "wb") as f:
            pickle.dump(planar_joint_angles, f)

        print("Planar joint angles in global coordinate frame are saved at : planar_global_joints_angles.pkl")

    def planar_angle_values_plot_pushButton_clicked(self):
        with open(ProjectWorkingPath + '/task_result_files/planar_global_joints_angles.pkl', "rb") as f_planar:
            planar_joint_angles = pickle.load(f_planar)

        with open(ProjectWorkingPath + '/task_result_files/absolute_joint_angles.pkl', "rb") as f_absolute:
            absolute_joint_angles = pickle.load(f_absolute)

        selected_angle = self.planar_angle_selector_comboBox.currentText()

        if selected_angle in absolute_joint_angles[0].keys():
            print(f"Plotting {selected_angle}")
            plt.figure()  # Create a new figure
            plt.title(selected_angle)  # Set the title to the selected angle
            plt.plot(absolute_joint_angles[0][selected_angle],'k', lw=3)
            plt.plot(planar_joint_angles[selected_angle][:,0],'r')
            plt.plot(planar_joint_angles[selected_angle][:,1],'g')
            plt.plot(planar_joint_angles[selected_angle][:,2],'b')

            plt.legend(['Absolute', 'X', 'Y', 'Z'])

            plt.show()

    def euler_calculate_pushButton_clicked(self):
        global relative_euler_angles_roll, relative_euler_angles_pitch, relative_euler_angles_yaw

        first_selected_segment_for_euler = self.first_segment_selector_euler_comboBox.currentText()
        second_selected_segment_for_euler = self.second_segment_selector_euler_comboBox.currentText()

        with open(ProjectWorkingPath + '/task_result_files/extrinsic_segment_orientations.pkl', "rb") as f:
            all_extrinsic_angles = pickle.load(f)

        relative_euler_angles_roll, relative_euler_angles_pitch, relative_euler_angles_yaw = \
            calculate_euler_angles_from_segments_orientations(first_selected_segment_for_euler, second_selected_segment_for_euler, all_extrinsic_angles)

        # Create subplots for each Euler angle component
        fig, axs = plt.subplots(3, 1, figsize=(8, 10))

        # Plot each Euler angle component in separate subplots
        axs[0].plot(relative_euler_angles_roll)
        # axs[0].set_xlabel('Frame Index')
        axs[0].set_ylabel('X-axis (Roll)')

        axs[1].plot(relative_euler_angles_pitch)
        # axs[1].set_xlabel('Frame Index')
        axs[1].set_ylabel('Y-axis (Pitch)')

        axs[2].plot(relative_euler_angles_yaw)
        axs[2].set_xlabel('Frame')
        axs[2].set_ylabel('Z-axis (Yaw)')

        plt.tight_layout()
        plt.show()

    def euler_angle_save_pushButton_clicked(self):
        # Create a new Workbook
        wb = openpyxl.Workbook()

        # Remove the default sheet that is created automatically
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Create a new sheet with a name based on the selected segments
        first_selected_segment_for_euler = self.first_segment_selector_euler_comboBox.currentText()
        second_selected_segment_for_euler = self.second_segment_selector_euler_comboBox.currentText()
        sheet_name = f"{first_selected_segment_for_euler}_{second_selected_segment_for_euler}"
        ws = wb.create_sheet(title=sheet_name)

        # Set the header row with direction names
        direction_names = ['X-axis (Roll)', 'Y-axis (Pitch)', 'Z-axis (Yaw)']
        ws.append(direction_names)

        # Append the Euler angles data to the sheet
        for roll, pitch, yaw in zip(relative_euler_angles_roll, relative_euler_angles_pitch, relative_euler_angles_yaw):
            ws.append([roll, pitch, yaw])

        # Save the workbook to a file
        file_path = os.path.join(ProjectWorkingPath, "task_result_files", f"{sheet_name}.xlsx")
        wb.save(file_path)
        print(
            f"Euler angles between {first_selected_segment_for_euler} and {second_selected_segment_for_euler} were calculated and saved to Excel.")

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    Form = QtWidgets.QWidget()
    ui = Ui_Form()
    ui.setupUi(Form)
    Form.show()
    sys.exit(app.exec_())
